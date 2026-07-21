#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
文件名: check_summary_person.py
作者: bao0
创建日期: 2026-07-16
描述: 校验接口返回 JSON 中 SUMMARY_PERSON 模板内容是否符合需求文档

需求规则（课后个性化总结-含阿语翻译.docx）:
  - 优秀表现（good_performance）: 每个【知识类型】取掌握度最高的知识点（相同则随机），
    掌握度 >= 3 时纳入，根据该知识点的【能力项】展示对应文案
  - 需要加强（need_strengthen）: 每个【知识类型】取掌握度最低的知识点（相同则随机），
    掌握度 <= 2 时纳入，根据该知识点的【能力项】展示对应文案

固定字段校验:
  - good_performance_title == "I can do it!"
  - need_strengthen_title  == "I need to do it!"
  - good_default  含 "继续加油，一起看看下面需要加强的内容"（无优秀表现时兜底）
  - need_default  含 "基础难度全部通过，让我们一起挑战高难度吧！"（无需加强时兜底）

各课型文案模板（need_strengthen 部分）:
  词汇句型课(101):
    词汇/词组 + 输出-说  → 再练习一下单词 "{word}"
    词汇/词组 + 输入-听  → 再练习一下单词 "{word}"
    词汇/词组 + 输入-读  → 重点练习单词 "{word}"
    句型     + 输入-听  → 再挑战一下句型 "{sentence}"
    句型     + 输出-说  → 再挑战一下句型 "{sentence}"
    句型     + 交际-说  → 还需要练习句型 "{sentence}"
  字母课(102):
    字母 + 字母-大小写匹配/命名/认读/自拼-认读 → 再练习一下字母 "{letter}"
    代表单词 + 输入-听/输出-说               → 还需要练习单词 "{word}"
  自拼课(103):
    自拼 + 自拼-认读    → 重点练习一下自拼规则 "{phonics}"
    代表单词 + 自拼-拼读/拼写/输入-听 → 还需要练习单词 "{word}"
  对话课(105):
    词汇词组 + 输出-说  → 再练习一下单词 "{word}"
    句型     + 输出-说  → 再挑战一下句型 "{sentence}"
    句型     + 交际-说  → 还需要练习句型 "{sentence}"
    功能句   + 交际-说  → 还需要练习功能句 "{sentence}"
  阅读课(100):
    句型 + 输入-听  → 再练习一下句型 "{sentence}"
    句型 + 输出-说  → 再挑战一下句型 "{sentence}"
    绘本 + 输入-读  → 再来读一下绘本 "{title}"

使用方式:
  python cocos/check_summary_person.py <json目录或单个json文件> [--output <输出目录>]
  python cocos/check_summary_person.py cocos/6种课型的接口返回json/
  python cocos/check_summary_person.py cocos/6种课型的接口返回json/101词汇not_mastered.json
"""

import json
import os
import sys
import glob
import re
import argparse
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment


# ==========================================
# 固定字段期望值
# ==========================================

EXPECTED_GOOD_TITLE = "I can do it!"
EXPECTED_NEED_TITLE = "I need to do it!"
EXPECTED_GOOD_DEFAULT = [
    "继续加油，一起看看下面需要加强的内容",
    "Keep it up! Let's see what we can work on next.",
]
EXPECTED_NEED_DEFAULT = [
    "基础难度全部通过，让我们一起挑战高难度吧！",
    "Basics cleared! Ready to level up?",
]

# ==========================================
# 各课型文案前缀规则
# knowledge_type（接口返回英文）→ 中文类型
# ==========================================

KNOWLEDGE_TYPE_MAP = {
    'sentence_patterns': '句型',
    'reading':                     '绘本',
    'words': '词汇',
    'words': '词汇词组',
    'letters': '字母',
    'letter_words': '字母代表词',
    #'reading': '歌谣',
    'phonics': '自拼',
    'phonics_words': '自拼代表词',
    'functional_language': '功能句',

    

    'word_sense': '词汇',    
    'phrase_sense': '词汇',
    'sentence': '句型',
    'sentence_pattern_structure': '句型',
    'alphabet_rule': '字母',
    'alphabet_rule_example_word': '字母代表词',
    'chant': '阅读',

    'phonics_rule': '自拼',
    'phonics_rule_example_word': '自拼代表词',
    'conversation': '对话',
    'conversation_sentence': '对话某轮中的某一句',
    'reading_content':             '绘本',
    'functional_sentence':         '功能句',
    'sentence_pattern_example_sentence': '句式例句'    
}

ABILITY_TYPE_MAP = {
    "INPUT-LISTEN":       "输入-听",
    "OUTPUT-SPEAK":       "输出-说",
    "INPUT-READ":         "输入-读",
    "INTERACTION-SPEAK":  "交际-说",
    "PHONICS-READ":       "自拼-认读",
    "PHONICS-SPELLING":   "自拼-拼写",
    "PHONICS-WRITE":      "自拼-拼写",
    "LETTER-MATCH":       "字母-大小写匹配",
    "LETTER-NAME":        "字母-命名",
    "LETTER-READ":        "字母-认读",    
}

# 文案前缀规则: (知识类型中文, 能力中文) → 前缀
# 前缀中 {word}/{sentence}/{letter}/{phonics}/{title} 为占位符，校验时做前缀匹配
NEED_STRENGTHEN_PREFIX_RULES = {
    # 词汇句型课 101
    ("词汇", "输入-听"):         ["再练习一下单词",       "Let's practice the word"],
    ("词汇", "输出-说"):         ["再练习一下单词",       "Let's practice the word"],
    ("词汇", "输入-读"):         ["重点练习单词",         "Let's practice the word"],
    ("句型", "输入-听"):         ["再挑战一下句型",       "Try the sentence again:"],
    ("句型", "输出-说"):         ["再练习一下句型",       "Try the sentence again:"],
    ("句型", "交际-说"):         ["还需要练习句型",       "Try to answer questions with"],
    # 字母课 102
    ("字母", "字母-大小写匹配"): ["再练习一下字母",       "Let's practice the letter"],
    ("字母", "字母-命名"):       ["再练习一下字母",       "Let's practice the letter"],
    ("字母", "字母-认读"):       ["重点练习一下字母",     "Let's practice the letter"],
    ("字母", "自拼-认读"):       ["重点练习一下字母",     "Let's practice the letter"],
    ("字母代表词", "输入-听"):   ["还需要练习单词",       "Try the word again:"],
    ("字母代表词", "输出-说"):   ["还需要练习单词",       "Try the word again:"],
    # 自拼课 103
    ("自拼", "自拼-认读"):       ["重点练习一下自拼规则", "Let's practice the phonics rule"],
    ("自拼代表词", "自拼-拼读"): ["还需要练习单词",       "Try the word again:"],
    ("自拼代表词", "自拼-拼写"): ["还需要练习单词",       "Try the word again:"],
    ("自拼代表词", "输入-听"):   ["还需要练习单词",       "Try the word again:"],
    # 对话课 105
    ("对话词汇", "输出-说"):     ["再练习一下单词",       "Let's practice the word"],
    ("对话句型", "输出-说"):     ["再挑战一下句型",       "Try the sentence again:"],
    ("对话句型", "交际-说"):     ["还需要练习句型",       "Try the sentence again:"],
    ("功能句", "交际-说"):       ["还需要练习功能句",     "Try the functional sentence again:"],
    # 阅读课 100
    ("句型", "输入-听"):         ["再练习一下句型",       "Try the sentence again:"],
    ("句型", "输出-说"):         ["再挑战一下句型",       "Try the sentence again:"],
    ("绘本", "输入-读"):         ["再来读一下绘本",       "Try to read the story again:"],
}

GOOD_PERFORMANCE_PREFIX_RULES = {
    # 词汇句型课 101
    ("词汇", "输入-听"):         ["能听懂",                  "Can understand the word"],
    ("词汇", "输出-说"):         ["能说出",                  "Can say the word"],
    ("词汇", "输入-读"):         ["能读出",                  "Can read the word"],
    ("句型", "输入-听"):         ["能听懂",                  "Can understand the sentence"],
    ("句型", "输出-说"):         ["能说出",                  "Can say the sentence"],
    ("句型", "交际-说"):         ["能在所学情境中使用",       "Can use the sentence in learned contexts:"],
    # 字母课 102
    ("字母", "字母-大小写匹配"): ["能匹配大小写",            "Can match big and small"],
    ("字母", "字母-命名"):       ["能命名字母",              "Can name the letter"],
    ("字母", "字母-认读"):       ["能认读字母",              "Can recognize the letter"],
    ("字母", "自拼-认读"):       ["能认出",                  "Can recognize the sound of"],
    ("字母代表词", "输入-听"):   ["能听懂单词",              "Can understand the word"],
    ("字母代表词", "输出-说"):   ["能说出单词",              "Can say the word"],
    # 自拼课 103
    ("自拼", "自拼-认读"):       ["能认读自拼规则",          "Can recognize the phonics rule"],
    ("自拼代表词", "自拼-拼读"): ["能拼读出单词",            "Can blend the word"],
    ("自拼代表词", "自拼-拼写"): ["能拼写出单词",            "Can spell the word"],
    ("自拼代表词", "输入-听"):   ["能听懂单词",              "Can understand the word"],
    # 对话课 105
    ("对话词汇", "输出-说"):     ["能说出",                  "Can say the word"],
    ("对话句型", "输出-说"):     ["能说出",                  "Can say the sentence"],
    ("对话句型", "交际-说"):     ["能在所学情境中使用",       "Can use the sentence in learned contexts:"],
    ("功能句", "交际-说"):       ["能在所学情境使用功能句",   "Can use the functional sentence in learned contexts:"],
    # 阅读课 100
    ("绘本", "输入-读"):         ["能读懂故事",              "Can understand the story:"],
}


# ==========================================
# Excel 报告
# ==========================================

class ExcelReport:
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "校验结果"

        self.ws.column_dimensions["A"].width = 28
        self.ws.column_dimensions["B"].width = 15
        self.ws.column_dimensions["C"].width = 30
        self.ws.column_dimensions["D"].width = 10
        self.ws.column_dimensions["E"].width = 60

        headers = ["测试文件", "检查项", "实际内容", "结果", "说明"]
        self.ws.append(headers)
        hfill = PatternFill(fill_type="solid", fgColor="4472C4")
        hfont = Font(color="FFFFFF", bold=True)
        for col in range(1, 6):
            cell = self.ws.cell(1, col)
            cell.fill = hfill
            cell.font = hfont
            cell.alignment = Alignment(horizontal="center")

        self.pass_fill = PatternFill(fill_type="solid", fgColor="C6EFCE")
        self.fail_fill = PatternFill(fill_type="solid", fgColor="FFC7CE")
        self.warn_fill = PatternFill(fill_type="solid", fgColor="FFEB9C")
        self.pass_count = self.fail_count = self.warn_count = 0

    def add(self, filename, item, actual, result, msg=""):
        self.ws.append([filename, item, str(actual) if actual is not None else "", result, msg])
        row = self.ws.max_row
        if result == "PASS":
            self.ws.cell(row, 4).fill = self.pass_fill
            self.pass_count += 1
        elif result == "WARN":
            self.ws.cell(row, 4).fill = self.warn_fill
            self.warn_count += 1
        else:
            self.ws.cell(row, 4).fill = self.fail_fill
            self.fail_count += 1

    def init_detail_sheet(self):
        """初始化知识点对比 sheet"""
        self.ws_detail = self.wb.create_sheet("知识点对比")
        headers = [
            "测试文件", "知识类型", "知识点名称", "能力项", "掌握度",
            "归属(good≥3/need≤2)",
            "预期good_performance", "实际good_performance",
            "预期need_strengthen", "实际need_strengthen",
            "对比结果"
        ]
        self.ws_detail.append(headers)
        hfill = PatternFill(fill_type="solid", fgColor="4472C4")
        hfont = Font(color="FFFFFF", bold=True)
        for col in range(1, len(headers) + 1):
            cell = self.ws_detail.cell(1, col)
            cell.fill = hfill
            cell.font = hfont
            cell.alignment = Alignment(horizontal="center")
        col_widths = [28, 12, 16, 14, 8, 18, 40, 40, 40, 40, 10]
        for i, w in enumerate(col_widths, 1):
            from openpyxl.utils import get_column_letter
            self.ws_detail.column_dimensions[get_column_letter(i)].width = w

    def add_detail_row(self, filename, kt, kname, ability, mastery,
                       group, exp_good, exp_need, act_good, act_need, match):
        row = [filename, kt, kname, ability, mastery, group,
               exp_good, act_good, exp_need, act_need, match]
        self.ws_detail.append(row)
        result_col = 11
        row_idx = self.ws_detail.max_row
        if match == "PASS":
            self.ws_detail.cell(row_idx, result_col).fill = self.pass_fill
        elif match == "WARN":
            self.ws_detail.cell(row_idx, result_col).fill = self.warn_fill
        else:
            self.ws_detail.cell(row_idx, result_col).fill = self.fail_fill

    def save(self, output_dir="."):
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = os.path.join(output_dir, f"check_summary_person_{ts}.xlsx")
        self.wb.save(path)
        return path


# ==========================================
# 检查器
# ==========================================

class Checker:
    def __init__(self, report: ExcelReport, filename: str):
        self.report = report
        self.filename = os.path.basename(filename)

    def ok(self, item, actual="", msg=""):
        self.report.add(self.filename, item, actual, "PASS", msg)
        print(f"  √ {item}")

    def fail(self, item, actual, msg):
        self.report.add(self.filename, item, actual, "FAIL", msg)
        print(f"  × {item} | 实际: {actual} | {msg}")

    def warn(self, item, actual, msg):
        self.report.add(self.filename, item, actual, "WARN", msg)
        print(f"  ! {item} | 实际: {actual} | {msg}")


# ==========================================
# 核心校验逻辑
# ==========================================

def _map_knowledge_type(raw: str) -> str:
    return KNOWLEDGE_TYPE_MAP.get(raw, raw or "")


def _map_ability_type(raw: str) -> str:
    return ABILITY_TYPE_MAP.get(raw, raw or "")


def _check_fixed_fields(content: dict, checker: Checker):
    """校验固定字段文案"""
    checks = [
        ("good_performance_title", EXPECTED_GOOD_TITLE),
        ("need_strengthen_title",  EXPECTED_NEED_TITLE),
    ]
    for field, expected in checks:
        actual = content.get(field)
        if actual == expected:
            checker.ok(f"{field}='{expected}'", actual)
        else:
            checker.fail(field, actual, f"期望: '{expected}'")

    # good_default / need_default 是列表，校验是否包含期望文案（中文或英文任意一条即PASS）
    for field, expected_texts in [
        ("good_default", EXPECTED_GOOD_DEFAULT),
        ("need_default", EXPECTED_NEED_DEFAULT),
    ]:
        val = content.get(field)
        if isinstance(val, list) and len(val) > 0:
            if any(e in val for e in expected_texts):
                checker.ok(f"{field} 含期望兜底文案", val)
            else:
                checker.fail(field, val, f"期望包含(任意一个): {expected_texts}")
        else:
            checker.fail(field, val, "列表为空或不是list类型")


def _infer_mastery(filename: str) -> str:
    name = filename.lower()
    if "not_mastered" in name:
        return "not_mastered"
    if "excellent" in name:
        return "excellent"
    if "moderate" in name:
        return "moderate"
    return "unknown"


def _check_performance_list(items: list, field: str, checker: Checker, mastery: str):
    """
    校验 good_performance / need_strengthen 列表：
    - not_mastered: good_performance 应为空，need_strengthen 应有数据
    - excellent: good_performance 应有数据，need_strengthen 应为空
    - moderate: 两者都可能有数据
    """
    good = items[0] if len(items) > 0 else None
    need = items[1] if len(items) > 1 else None
    # 此函数只负责数量符合预期，文案前缀在 _check_text_prefix 中处理

    if field == "good_performance":
        val = good
        if mastery == "not_mastered":
            if val == [] or val is None:
                checker.ok("good_performance 为空（not_mastered 场景正确）", val)
            else:
                checker.warn("good_performance", val, "not_mastered 场景预期为空，实际有数据")
        elif mastery == "excellent":
            if val and len(val) > 0:
                checker.ok(f"good_performance 有数据（共{len(val)}条）", val)
            else:
                checker.fail("good_performance", val, "excellent 场景应有数据")

    elif field == "need_strengthen":
        val = need
        if mastery == "not_mastered":
            if val and len(val) > 0:
                checker.ok(f"need_strengthen 有数据（共{len(val)}条）", val)
            else:
                checker.fail("need_strengthen", val, "not_mastered 场景应有数据")
        elif mastery == "excellent":
            if val == [] or val is None:
                checker.ok("need_strengthen 为空（excellent 场景正确）", val)
            else:
                checker.warn("need_strengthen", val, "excellent 场景预期为空，实际有数据")


def _check_text_prefix(text_list: list, field: str, checker: Checker):
    """
    校验 good_performance / need_strengthen 中每条文案的前缀是否符合需求。
    文案格式：前缀 "{知识点名称}"
    通过正则提取前缀，对照规则表校验。
    """
    if not text_list:
        return

    rules = NEED_STRENGTHEN_PREFIX_RULES if field == "need_strengthen" else GOOD_PERFORMANCE_PREFIX_RULES

    # 收集所有已知前缀，按长度降序（优先匹配最长前缀，避免短前缀误命中）
    all_prefixes_sorted = sorted(
        set(item for v in rules.values() for item in (v if isinstance(v, list) else [v])),
        key=len, reverse=True
    )

    for text in text_list:
        # 格式：前缀 + 空格 + 知识点名称，例如 "再练习一下单词 spider"
        matched_prefix = next(
            (p for p in all_prefixes_sorted if text.startswith(p + ' ') or text == p),
            None
        )

        if matched_prefix is not None:
            matched_rules = [
                (kt, ab) for (kt, ab), p in rules.items()
                if (isinstance(p, list) and matched_prefix in p) or p == matched_prefix
            ]
            checker.ok(
                f"{field} 文案前缀",
                text,
                f"前缀'{matched_prefix}'匹配规则: {matched_rules}"
            )
        else:
            similar = [p for p in all_prefixes_sorted if text[:4] in p or p[:4] in text]
            hint = f"相似前缀: {similar}" if similar else f"所有合法前缀: {sorted(all_prefixes_sorted)}"
            checker.fail(
                f"{field} 文案前缀",
                text,
                f"前缀不在规则表中。{hint}"
            )


def _extract_knowledge_points(res: list) -> list:
    """
    从 SUMMARY_V1 / SUMMARY_V2 的 knowledges 中提取所有知识点。
    返回: [{"knowledge_type": str, "knowledge_name": str, "ability": str, "mastery": int}, ...]
    每个知识点按 ability_level 里的每个能力项展开为独立行。
    """
    points = []
    for item in res:
        code = item.get("template_code")
        if code not in ("SUMMARY_V1", "SUMMARY_V2"):
            continue
        knowledges = (item.get("data") or {}).get("knowledges") or {}
        for key, value in knowledges.items():
            if key == "video" or not value:
                continue
            entries = value if isinstance(value, list) else [value]
            for entry in entries:
                if not isinstance(entry, dict):
                    continue
                kt = entry.get("knowledge_type", "")
                kname = entry.get("knowledge_name", "")
                ability_level = entry.get("ability_level") or {}
                for ability, mastery in ability_level.items():
                    points.append({
                        "knowledge_type": kt,
                        "knowledge_name": kname,
                        "ability": ability,
                        "mastery": mastery,
                    })
    return points


def _build_candidate_texts(kname: str, mastery: int, pts_at_mastery: list, rules: dict, kt_cn: str) -> tuple:
    """
    对同一知识点下掌握度相同的多个能力项，生成所有候选文案。
    返回: (candidate_texts, display_text)
      candidate_texts: 所有候选文案列表（用于校验时任意匹配一条即PASS）
      display_text: 预期展示文案（单条直接显示，多条注明"取任何一个都正确"）
    """
    candidates = []
    for p in pts_at_mastery:
        ability_cn = ABILITY_TYPE_MAP.get(p["ability"], p["ability"])
        prefix = rules.get((kt_cn, ability_cn))
        text = f'{prefix} "{kname}"' if prefix else f'[未知规则:{kt_cn}/{ability_cn}] "{kname}"'
        if text not in candidates:
            candidates.append(text)

    if len(candidates) == 1:
        display = candidates[0]
    else:
        display = " 或 ".join(f'"{t}"' for t in candidates) + "（取任何一个都正确）"

    return candidates, display


def _calc_expected_good_need(points: list) -> tuple:
    """
    按需求规则推算预期的 good_performance 和 need_strengthen 文案列表。

    规则：
      - good_performance：每个【知识点（knowledge_name）】取掌握度最高的知识（相同则随机），
        根据该知识点的【能力项】展示对应文案；掌握度 >= 3 时才纳入 good_performance；
        若同一知识点下多个能力项掌握度相同，则列出所有候选文案并标注"取任何一个都正确"
      - need_strengthen：每个【知识点（knowledge_name）】取掌握度最低的知识（相同则随机），
        规则同上；掌握度 <= 2 时才纳入 need_strengthen

    返回:
      exp_good: 预期文案列表（每条可能是多候选）
      exp_need: 预期文案列表
      selected: [(kt_cn, ability_cn, kname, mastery, display_text, group, candidates), ...]
    """
    from collections import defaultdict

    # 按 knowledge_type 分组——每个知识类型只取一个代表知识点
    groups_by_kt = defaultdict(list)
    for p in points:
        groups_by_kt[p["knowledge_type"]].append(p)

    exp_good = []
    exp_need = []
    selected = []

    for kt, pts in groups_by_kt.items():
        kt_cn = KNOWLEDGE_TYPE_MAP.get(kt, kt)

        # ── good_performance：取掌握度最高的知识点
        max_mastery = max(p["mastery"] for p in pts)
        if max_mastery >= 3:
            pts_max = [p for p in pts if p["mastery"] == max_mastery]
            # 掌握度相同时，每个知识点+能力项组合都是候选
            candidates_good = []
            for p in pts_max:
                ability_cn = ABILITY_TYPE_MAP.get(p["ability"], p["ability"])
                prefixes = GOOD_PERFORMANCE_PREFIX_RULES.get((kt_cn, ability_cn))
                if prefixes is None:
                    text = f'[未知规则:{kt_cn}/{ability_cn}] {p["knowledge_name"]}'
                    if text not in candidates_good:
                        candidates_good.append(text)
                else:
                    for pref in (prefixes if isinstance(prefixes, list) else [prefixes]):
                        text = f'{pref} {p["knowledge_name"]}'
                        if text not in candidates_good:
                            candidates_good.append(text)
            if len(candidates_good) == 1:
                display_good = candidates_good[0]
            else:
                display_good = " 或 ".join(f'"{t}"' for t in candidates_good) + "（取任何一个都正确）"
            exp_good.extend(candidates_good)
            ability_cn_good = ABILITY_TYPE_MAP.get(pts_max[0]["ability"], pts_max[0]["ability"])
            kname_good = pts_max[0]["knowledge_name"]
            selected.append((kt_cn, ability_cn_good, kname_good, max_mastery, display_good, "good", candidates_good))

        # ── need_strengthen：取掌握度最低的知识点
        min_mastery = min(p["mastery"] for p in pts)
        if min_mastery <= 2:
            pts_min = [p for p in pts if p["mastery"] == min_mastery]
            candidates_need = []
            for p in pts_min:
                ability_cn = ABILITY_TYPE_MAP.get(p["ability"], p["ability"])
                prefixes = NEED_STRENGTHEN_PREFIX_RULES.get((kt_cn, ability_cn))
                if prefixes is None:
                    text = f'[未知规则:{kt_cn}/{ability_cn}] {p["knowledge_name"]}'
                    if text not in candidates_need:
                        candidates_need.append(text)
                else:
                    for pref in (prefixes if isinstance(prefixes, list) else [prefixes]):
                        text = f'{pref} {p["knowledge_name"]}'
                        if text not in candidates_need:
                            candidates_need.append(text)
            if len(candidates_need) == 1:
                display_need = candidates_need[0]
            else:
                display_need = " 或 ".join(f'"{t}"' for t in candidates_need) + "（取任何一个都正确）"
            exp_need.extend(candidates_need)
            ability_cn_need = ABILITY_TYPE_MAP.get(pts_min[0]["ability"], pts_min[0]["ability"])
            kname_need = pts_min[0]["knowledge_name"]
            selected.append((kt_cn, ability_cn_need, kname_need, min_mastery, display_need, "need", candidates_need))

    return exp_good, exp_need, selected


def validate_file(filepath: str, report: ExcelReport):
    print(f"\n{'=' * 60}")
    print(f"文件: {os.path.basename(filepath)}")
    print("=" * 60)

    try:
        with open(filepath, encoding="utf-8") as f:
            data = json.load(f)
    except Exception as e:
        print(f"× 文件读取失败: {e}")
        report.add(os.path.basename(filepath), "文件读取", "", "FAIL", str(e))
        return

    checker = Checker(report, filepath)
    mastery = _infer_mastery(filepath)
    print(f"推断掌握度场景: {mastery}")

    res = data.get("res", [])
    template_codes = {item.get("template_code") for item in res}

    # ── 嘉年华课型（104）──
    # 特征：有 SUMMARY_V2，无 SUMMARY_PERSON
    # 规则：嘉年华只展示本节课练习的知识点数量，不校验掌握度，不需要个性化总结
    if "SUMMARY_V2" in template_codes and "SUMMARY_PERSON" not in template_codes:
        print("  检测到嘉年华课型（SUMMARY_V2 无 SUMMARY_PERSON），跳过个性化总结校验")
        checker.ok("嘉年华课型确认", "", "无 SUMMARY_PERSON，符合需求（嘉年华不需要个性化总结）")
        if "PRE_POST_SETTLE" in template_codes:
            checker.ok("PRE_POST_SETTLE 存在")
        else:
            checker.fail("PRE_POST_SETTLE", "", "嘉年华课型必须有 PRE_POST_SETTLE")
        return

    # ── 全掌握场景（所有知识点掌握度=4）──
    # 特征：有 SUMMARY_PERFECT，无 SUMMARY_PERSON
    # 校验规则：SUMMARY_PERFECT 有 video；PRE_POST_SETTLE 存在
    if "SUMMARY_PERFECT" in template_codes:
        print("  检测到 SUMMARY_PERFECT，进入全掌握场景校验")

        # 校验 SUMMARY_PERFECT.data.content.video 或 knowledges.video
        perfect_item = next(i for i in res if i.get("template_code") == "SUMMARY_PERFECT")
        data_obj = perfect_item.get("data") or {}
        video = (data_obj.get("knowledges") or {}).get("video") or \
                (data_obj.get("content") or {}).get("video")
        if video:
            checker.ok("SUMMARY_PERFECT.video 存在", video)
        else:
            checker.fail("SUMMARY_PERFECT.video", video, "全掌握场景 video 不能为空")

        # 校验 PRE_POST_SETTLE 存在
        if "PRE_POST_SETTLE" in template_codes:
            checker.ok("PRE_POST_SETTLE 存在")
        else:
            checker.fail("PRE_POST_SETTLE", "", "全掌握场景必须有 PRE_POST_SETTLE")

        # 全掌握场景不应有 SUMMARY_PERSON
        if "SUMMARY_PERSON" in template_codes:
            checker.warn("SUMMARY_PERSON", "", "全掌握场景不应返回 SUMMARY_PERSON")
        return

    # ── 普通场景：校验 SUMMARY_PERSON ──
    person_items = [i for i in res if i.get("template_code") == "SUMMARY_PERSON"]

    if not person_items:
        checker.warn("SUMMARY_PERSON", "", "该文件中无 SUMMARY_PERSON 模板，跳过")
        return

    for person in person_items:
        content = (person.get("data") or {}).get("content") or {}

        # 1. 固定字段校验
        _check_fixed_fields(content, checker)

        # 2. good_performance / need_strengthen 数量符合掌握度场景
        good = content.get("good_performance")
        need = content.get("need_strengthen")

        if mastery != "unknown":
            _check_performance_list([good, need], "good_performance", checker, mastery)
            _check_performance_list([good, need], "need_strengthen", checker, mastery)
        else:
            checker.warn("掌握度场景", "", "文件名无法推断掌握度，跳过数量校验")

        # 3. 文案前缀校验
        if isinstance(good, list) and good:
            _check_text_prefix(good, "good_performance", checker)
        if isinstance(need, list) and need:
            _check_text_prefix(need, "need_strengthen", checker)

        # 4. audio 字段存在性
        audio = content.get("audio")
        if audio:
            checker.ok("audio 字段存在", audio)
        else:
            checker.warn("audio 字段", audio, "audio 为空，请确认是否正常")

        # 5. 知识点对比：每个【知识类型-能力项】取掌握度最低的知识点，对比预期与实际
        if not hasattr(report, 'ws_detail'):
            report.init_detail_sheet()

        points = _extract_knowledge_points(res)
        exp_good, exp_need, selected = _calc_expected_good_need(points)
        act_good = good if isinstance(good, list) else []
        act_need = need if isinstance(need, list) else []
        filename_base = os.path.basename(filepath)

        print(f"\n  {'─' * 56}")
        print(f"  知识点对比（good=每【知识类型】取最高掌握度，need=每【知识类型】取最低掌握度）")
        print(f"  {'─' * 56}")
        print(f"  {'知识类型':<10} {'能力项':<16} {'知识点':<20} {'掌握度':<6} {'归属':<8} {'对比'}")
        print(f"  {'─' * 56}")

        # 预计算每行的匹配结果，以便后续排除已命中条目
        def _compute_matched(candidates, act_list):
            matched = [t for t in act_list if t in candidates]
            if not matched:
                cand_knames = set()
                for c in candidates:
                    clean = re.sub(r'["""\u201c\u201d]', '', c)
                    cand_knames.add(clean.split()[-1] if ' ' in clean else clean)
                if not cand_knames:
                    cand_knames = {candidates[0].split()[-1] if candidates else ""}
                matched = [t for t in act_list if any(kn in t for kn in cand_knames)]
            return matched

        rows_data = []
        for entry in selected:
            kt_cn, ability_cn, kname, mastery, display_text, group, candidates = entry
            act_list = act_good if group == "good" else act_need
            act_matched = _compute_matched(candidates, act_list)
            rows_data.append((kt_cn, ability_cn, kname, mastery, display_text, group, candidates, act_list, act_matched))

        # 汇总所有已命中条目
        globally_matched_good = set(t for *_, group, _, act_list, act_matched in rows_data if group == "good" for t in act_matched)
        globally_matched_need = set(t for *_, group, _, act_list, act_matched in rows_data if group == "need" for t in act_matched)

        for kt_cn, ability_cn, kname, mastery, display_text, group, candidates, act_list, act_matched in rows_data:
            group_label = "good(≥3)" if group == "good" else "need(≤2)"

            if act_matched:
                match = "PASS"
                match_sym = "√"
                act_str = "\n".join(act_matched)
            else:
                match = "FAIL" if act_list else "WARN"
                match_sym = "×" if match == "FAIL" else "!"
                # 展示未被其他行命中的实际条目
                globally_matched = globally_matched_good if group == "good" else globally_matched_need
                act_str = "\n".join(t for t in act_list if t not in globally_matched)

            print(f"  {kt_cn:<10} {ability_cn:<16} {kname:<20} {mastery:<6} {group_label:<8} {match_sym}")
            print(f"    预期: {display_text}")
            if act_str:
                for line in act_str.split("\n"):
                    print(f"    实际: {line}")
            else:
                print(f"    实际: （列表为空）")

            exp_good_str = display_text if group == "good" else ""
            exp_need_str = display_text if group == "need" else ""
            act_good_str = act_str if group == "good" else ""
            act_need_str = act_str if group == "need" else ""

            report.add_detail_row(
                filename_base, kt_cn, kname, ability_cn, mastery,
                group_label, exp_good_str, exp_need_str,
                act_good_str, act_need_str, match
            )

        print(f"  {'─' * 56}")
        print(f"  预期 good_performance: {exp_good}")        
        print(f"  实际 good_performance: {act_good}")
        print(f"  预期 need_strengthen : {exp_need}")
        print(f"  实际 need_strengthen : {act_need}")


# ==========================================
# 主函数
# ==========================================

def main():
    parser = argparse.ArgumentParser(description="SUMMARY_PERSON 文案校验工具")
    parser.add_argument("target", help="JSON文件路径或目录")
    parser.add_argument("--output", default=".", help="报告输出目录（默认当前目录）")
    args = parser.parse_args()

    report = ExcelReport()
    report.init_detail_sheet()

    target = args.target
    files = []
    if os.path.isdir(target):
        files = sorted(glob.glob(os.path.join(target, "*.json")))
    elif os.path.isfile(target):
        files = [target]
    else:
        files = sorted(glob.glob(target))

    if not files:
        print(f"未找到匹配文件: {target}")
        sys.exit(1)

    print(f"共找到 {len(files)} 个JSON文件")
    for f in files:
        validate_file(f, report)

    report_path = report.save(args.output)

    print(f"\n{'=' * 60}")
    print(f"校验完成")
    print(f"  PASS : {report.pass_count}")
    print(f"  FAIL : {report.fail_count}")
    print(f"  WARN : {report.warn_count}")
    print(f"  报告 : {report_path}")
    print("=" * 60)

    sys.exit(1 if report.fail_count > 0 else 0)


if __name__ == "__main__":
    main()
