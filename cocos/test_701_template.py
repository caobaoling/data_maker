#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
文件名: test_701_template.py
作者: bao0
创建日期: 2026-07-14
描述: 701课后模版接口返回JSON数据测试脚本

测试用例覆盖:
  TC-01: SUMMARY_PERFECT - 全掌握场景的video字段校验
  TC-02: SUMMARY_V1 - 词汇/自拼课型knowledges结构校验
  TC-03: SUMMARY_V2 - 嘉年华多类知识点校验
  TC-04: SUMMARY_PERSON - 个性化汇总内容校验
  TC-05: PRE_POST_SETTLE - 结尾帧data必须为null
  TC-06: 整体响应结构校验
  TC-07: 掌握度与内容对应校验

使用方式:
  # 测试单个文件
  python test_701_template.py path/to/xxx.json

  # 测试目录下所有json文件
  python test_701_template.py 6种课型的接口返回json/

  # 指定掌握度参数
  python test_701_template.py xxx.json --mastery excellent
"""

import json
import os
import sys
import glob
import argparse
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment


# ==========================================
# 常量
# ==========================================

VALID_ABILITY_LEVELS = {1, 2, 3, 4}

# SUMMARY_PERSON的标题固定值（来自需求文档）
EXPECTED_TITLES = {
    "title": "My Learning Highlights",
    "good_performance_title": "I can do it!",
    "need_strengthen_title": "I need to do it!",
}

# good_performance对应的掌握度（掌握度>=3）
GOOD_MASTERY_LEVELS = {3, 4}
# need_strengthen对应的掌握度（掌握度<=2）
NEED_MASTERY_LEVELS = {1, 2}

# 全掌握课型（只有SUMMARY_PERFECT + PRE_POST_SETTLE）
PERFECT_ONLY_COURSE_TYPES = {"100阅读", "102字母", "105对话"}


# ==========================================
# Excel报告
# ==========================================

class ExcelReport:

    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "测试结果"

        # 设置列宽
        self.ws.column_dimensions["A"].width = 30
        self.ws.column_dimensions["B"].width = 20
        self.ws.column_dimensions["C"].width = 35
        self.ws.column_dimensions["D"].width = 10
        self.ws.column_dimensions["E"].width = 50

        # 表头
        headers = ["测试文件", "Template", "检查项", "结果", "说明"]
        self.ws.append(headers)
        header_fill = PatternFill(fill_type="solid", fgColor="4472C4")
        header_font = Font(color="FFFFFF", bold=True)
        for col in range(1, 6):
            cell = self.ws.cell(1, col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        self.pass_fill = PatternFill(fill_type="solid", fgColor="C6EFCE")
        self.fail_fill = PatternFill(fill_type="solid", fgColor="FFC7CE")
        self.warn_fill = PatternFill(fill_type="solid", fgColor="FFEB9C")

        self.pass_count = 0
        self.fail_count = 0
        self.warn_count = 0

    def add(self, filename, template, item, result, msg=""):
        self.ws.append([filename, template, item, result, msg])
        row = self.ws.max_row
        if result == "PASS":
            self.ws.cell(row, 4).fill = self.pass_fill
            self.pass_count += 1
        elif result == "WARN":
            self.ws.cell(row, 4).fill = self.warn_fill
            self.warn_count += 1
        else:
            self.ws.cell(row, 4).fill = self.fail_fill
            self.fail_count += 1

    def save(self, output_dir="."):
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = os.path.join(output_dir, f"test_701_result_{ts}.xlsx")
        self.wb.save(filename)
        return filename


# ==========================================
# 检查器
# ==========================================

class Checker:

    def __init__(self, report: ExcelReport, filename: str):
        self.report = report
        self.filename = os.path.basename(filename)

    def ok(self, template, item, msg=""):
        self.report.add(self.filename, template, item, "PASS", msg)
        print(f"  √ [{template}] {item}")

    def fail(self, template, item, msg):
        self.report.add(self.filename, template, item, "FAIL", msg)
        print(f"  × [{template}] {item} => {msg}")

    def warn(self, template, item, msg):
        self.report.add(self.filename, template, item, "WARN", msg)
        print(f"  ! [{template}] {item} => {msg}")

    def check_exist(self, value, template, item):
        """检查值不为None且不为空字符串"""
        if value is None:
            self.fail(template, item, "值为 null")
            return False
        if value == "":
            self.fail(template, item, "值为空字符串")
            return False
        self.ok(template, item)
        return True

    def check_field_exists(self, obj, field, template, item):
        """检查字段在对象中存在（值可以为空）"""
        if field not in obj:
            self.fail(template, item, f"字段 '{field}' 不存在")
            return False
        self.ok(template, item)
        return True

    def check_list(self, value, template, item):
        """检查值是非空列表"""
        if not isinstance(value, list):
            self.fail(template, item, f"不是list，实际类型: {type(value).__name__}")
            return False
        if len(value) == 0:
            self.fail(template, item, "列表为空")
            return False
        self.ok(template, item)
        return True

    def check_url(self, value, template, item):
        """检查值是有效URL"""
        if not value:
            self.fail(template, item, "URL为空")
            return False
        if not str(value).startswith("http"):
            self.fail(template, item, f"不是有效URL: {value[:50]}")
            return False
        self.ok(template, item)
        return True

    def check_null(self, value, template, item):
        """检查值必须为null"""
        if value is not None:
            self.fail(template, item, f"应为null，实际为: {str(value)[:50]}")
            return False
        self.ok(template, item)
        return True

    def check_in_range(self, value, allowed, template, item):
        """检查值在允许范围内"""
        if value not in allowed:
            self.fail(template, item, f"值 {value} 不在允许范围 {allowed}")
            return False
        self.ok(template, item)
        return True


# ==========================================
# TC-06: 整体响应结构校验
# ==========================================

class ResponseStructureValidator:

    def __init__(self, checker: Checker):
        self.c = checker

    def validate(self, data: dict):
        name = "RESPONSE"

        # code == "10000"
        if data.get("code") == "10000":
            self.c.ok(name, "code=10000")
        else:
            self.c.fail(name, "code=10000", f"实际值: {data.get('code')}")

        # message == "success"
        if data.get("message") == "success":
            self.c.ok(name, "message=success")
        else:
            self.c.fail(name, "message=success", f"实际值: {data.get('message')}")

        # res是非空列表
        res = data.get("res")
        if not self.c.check_list(res, name, "res非空列表"):
            return False

        # PRE_POST_SETTLE必须是最后一个
        last_code = res[-1].get("template_code")
        if last_code == "PRE_POST_SETTLE":
            self.c.ok(name, "PRE_POST_SETTLE在最后")
        else:
            self.c.fail(name, "PRE_POST_SETTLE在最后", f"最后一个是: {last_code}")

        return True


# ==========================================
# TC-01: SUMMARY_PERFECT
# ==========================================

class SummaryPerfectValidator:
    """
    需求: 全掌握场景（100阅读/102字母/105对话）
    仅包含视频，title/good_performance/need_strengthen等应为null
    """

    def __init__(self, checker: Checker):
        self.c = checker

    def validate(self, template: dict):
        name = "SUMMARY_PERFECT"
        data = template.get("data") or {}
        content = data.get("content") or {}

        # video必须有效
        self.c.check_url(content.get("video"), name, "content.video")

        # 以下字段应为null（不应有个性化内容）
        for field in ["title", "good_performance_title", "need_strengthen_title",
                      "good_performance", "need_strengthen",
                      "good_default", "need_default", "audio"]:
            self.c.check_null(content.get(field), name, f"content.{field}==null")

        # knowledges应为null（不含知识点数据）
        self.c.check_null(data.get("knowledges"), name, "knowledges==null")


# ==========================================
# TC-02: SUMMARY_V1
# ==========================================

class SummaryV1Validator:
    """
    需求: 101词汇/103自拼课型
    - 词汇课: words非空，含image和ability_level；有sentence_patterns
    - 自拼课: words为空，有phonics和phonics_words；ability_level允许PHONICS-*类型
    ability_level值必须在[1,2,3,4]
    """

    # 自拼课允许的ability类型（不在标准4种之内，但合法）
    PHONICS_ABILITY_KEYS = {
        "PHONICS-READ", "PHONICS-SPELLING", "PHONICS-WRITE",
        "INPUT-LISTEN", "OUTPUT-SPEAK"
    }

    def __init__(self, checker: Checker):
        self.c = checker

    def _is_phonics_course(self, knowledges: dict) -> bool:
        """通过phonics字段判断是否为自拼课"""
        phonics = knowledges.get("phonics")
        return bool(phonics and len(phonics) > 0)

    def validate(self, template: dict):
        name = "SUMMARY_V1"
        data = template.get("data") or {}
        knowledges = data.get("knowledges")

        if not knowledges:
            self.c.fail(name, "knowledges存在", "knowledges为空或null")
            return

        self.c.ok(name, "knowledges存在")

        # video
        self.c.check_url(knowledges.get("video"), name, "knowledges.video")

        is_phonics = self._is_phonics_course(knowledges)

        if is_phonics:
            # ---- 自拼课 (103) ----
            self.c.ok(name, "识别为自拼课型(phonics字段存在)")

            # phonics列表校验
            phonics = knowledges.get("phonics")
            for i, ph in enumerate(phonics):
                p = f"phonics[{i}]({ph.get('knowledge_name', '?')})"
                self.c.check_exist(ph.get("knowledge_name"), name, f"{p}.knowledge_name")
                self.c.check_url(ph.get("image"), name, f"{p}.image")
                ability = ph.get("ability_level")
                if not ability:
                    self.c.fail(name, f"{p}.ability_level", "ability_level为空")
                else:
                    self.c.ok(name, f"{p}.ability_level存在")
                    for k, v in ability.items():
                        self.c.check_in_range(v, VALID_ABILITY_LEVELS, name, f"{p}.{k}")

            # phonics_words校验
            phonics_words = knowledges.get("phonics_words")
            if self.c.check_list(phonics_words, name, "knowledges.phonics_words非空"):
                for i, pw in enumerate(phonics_words):
                    p = f"phonics_words[{i}]({pw.get('knowledge_name', '?')})"
                    self.c.check_exist(pw.get("knowledge_name"), name, f"{p}.knowledge_name")
                    self.c.check_url(pw.get("image"), name, f"{p}.image")
                    ability = pw.get("ability_level")
                    if not ability:
                        self.c.fail(name, f"{p}.ability_level", "ability_level为空")
                    else:
                        self.c.ok(name, f"{p}.ability_level存在")
                        for k, v in ability.items():
                            self.c.check_in_range(v, VALID_ABILITY_LEVELS, name, f"{p}.{k}")

        else:
            # ---- 词汇课 (101) ----
            self.c.ok(name, "识别为词汇课型")

            words = knowledges.get("words")
            if self.c.check_list(words, name, "knowledges.words非空"):
                for i, word in enumerate(words):
                    p = f"words[{i}]({word.get('knowledge_name', '?')})"
                    self.c.check_exist(word.get("knowledge_name"), name, f"{p}.knowledge_name")
                    self.c.check_url(word.get("image"), name, f"{p}.image")

                    ability = word.get("ability_level")
                    if not ability:
                        self.c.fail(name, f"{p}.ability_level", "ability_level为空")
                    else:
                        self.c.ok(name, f"{p}.ability_level存在")
                        for ability_key, ability_val in ability.items():
                            self.c.check_in_range(
                                ability_val, VALID_ABILITY_LEVELS,
                                name, f"{p}.ability_level.{ability_key}"
                            )

            # sentence_patterns（词汇课必须有）
            patterns = knowledges.get("sentence_patterns")
            if patterns:
                self.c.ok(name, "knowledges.sentence_patterns存在")
                for i, sp in enumerate(patterns):
                    p = f"sentence_patterns[{i}]({sp.get('knowledge_name', '?')})"
                    self.c.check_exist(sp.get("knowledge_name"), name, f"{p}.knowledge_name")
                    ability = sp.get("ability_level")
                    if not ability:
                        self.c.fail(name, f"{p}.ability_level", "ability_level为空")
                    else:
                        self.c.ok(name, f"{p}.ability_level存在")
                        for k, v in ability.items():
                            self.c.check_in_range(v, VALID_ABILITY_LEVELS, name, f"{p}.{k}")
            else:
                self.c.fail(name, "knowledges.sentence_patterns", "词汇课sentence_patterns不应为空")


# ==========================================
# TC-03: SUMMARY_V2
# ==========================================

class SummaryV2Validator:
    """
    需求: 104嘉年华课型
    按课程level不同展示不同标签:
      Lv0: Letters, Words of Letters, Words, Sentence Patterns, Topic
      Lv1~2: Phonics, Words of Phonics, Words, Sentence Patterns, Topic
    """

    def __init__(self, checker: Checker):
        self.c = checker

    def validate(self, template: dict):
        name = "SUMMARY_V2"
        data = template.get("data") or {}
        knowledges = data.get("knowledges")

        if not knowledges:
            self.c.fail(name, "knowledges存在", "knowledges为空或null")
            return

        self.c.ok(name, "knowledges存在")

        # video必须有效
        self.c.check_url(knowledges.get("video"), name, "knowledges.video")

        # words（所有level都要有）
        words = knowledges.get("words")
        if words is not None and len(words) > 0:
            self.c.ok(name, "knowledges.words存在")
            for i, word in enumerate(words):
                p = f"words[{i}]({word.get('knowledge_name', '?')})"
                self.c.check_exist(word.get("knowledge_id"), name, f"{p}.knowledge_id")
                self.c.check_exist(word.get("knowledge_name"), name, f"{p}.knowledge_name")
        else:
            self.c.warn(name, "knowledges.words", "words为空")

        # 判断是Lv0还是Lv1~2（根据有无phonics字段）
        phonics = knowledges.get("phonics")
        letters = knowledges.get("letters")

        if phonics and len(phonics) > 0:
            # Lv1~2: 应有phonics和phonics_words
            self.c.ok(name, "Lv1~2模式: phonics存在")
            phonics_words = knowledges.get("phonics_words")
            if phonics_words and len(phonics_words) > 0:
                self.c.ok(name, "Lv1~2模式: phonics_words存在")
            else:
                self.c.warn(name, "Lv1~2模式: phonics_words", "phonics_words为空")
        elif letters and len(letters) > 0:
            # Lv0: 应有letters和letters_words
            self.c.ok(name, "Lv0模式: letters存在")
            letters_words = knowledges.get("letters_words") or knowledges.get("letter_words")
            if letters_words and len(letters_words) > 0:
                self.c.ok(name, "Lv0模式: letters_words存在")
            else:
                self.c.warn(name, "Lv0模式: letters_words", "letters_words为空")
        else:
            self.c.warn(name, "课程level判断", "phonics和letters均为空，无法判断level")

        # topic（对话页图文）
        topic = knowledges.get("topic")
        if topic and len(topic) > 0:
            self.c.ok(name, "knowledges.topic存在")
            for i, t in enumerate(topic):
                p = f"topic[{i}]({t.get('knowledge_name', '?')})"
                self.c.check_exist(t.get("knowledge_id"), name, f"{p}.knowledge_id")
        else:
            self.c.warn(name, "knowledges.topic", "topic为空（可能是该课型不含对话）")

        # content应为null（V2不用content字段）
        self.c.check_null(data.get("content"), name, "content==null")


# ==========================================
# TC-04: SUMMARY_PERSON
# ==========================================

class SummaryPersonValidator:
    """
    需求: 个性化汇总模板
    - title固定为"My Learning Highlights"
    - good/need标题固定
    - good_performance: 掌握度3-4的知识点
    - need_strengthen: 掌握度1-2的知识点
    - good_default/need_default: 默认文案，不为空
    - audio字段必须存在（可为空字符串）
    """

    def __init__(self, checker: Checker):
        self.c = checker

    def validate(self, template: dict):
        name = "SUMMARY_PERSON"
        data = template.get("data") or {}
        content = data.get("content")

        if not content:
            self.c.fail(name, "content存在", "content为空或null")
            return

        self.c.ok(name, "content存在")

        # 固定标题校验
        for field, expected in EXPECTED_TITLES.items():
            actual = content.get(field)
            if actual == expected:
                self.c.ok(name, f"content.{field}=='{expected}'")
            else:
                self.c.fail(
                    name,
                    f"content.{field}",
                    f"期望: '{expected}'，实际: '{actual}'"
                )

        # audio字段必须存在（字段本身必须在content中）
        if "audio" in content:
            self.c.ok(name, "content.audio字段存在")
        else:
            self.c.fail(name, "content.audio字段存在", "audio字段缺失")

        # good_performance 和 need_strengthen：两者至少有一个有数据，或都为空列表/null
        # 接受：list（含空list）或 null
        # 不接受：非list类型
        good = content.get("good_performance")
        need = content.get("need_strengthen")

        if good is None and need is None:
            self.c.warn(
                name,
                "good_performance/need_strengthen",
                "两者均为null（全掌握场景可接受，请人工确认）"
            )
        else:
            if good is not None:
                if isinstance(good, list):
                    cnt = len(good)
                    if cnt > 0:
                        self.c.ok(name, f"content.good_performance (共{cnt}条)")
                    else:
                        # 空列表：说明该学生该知识点全部需要加强，无优秀表现，正常
                        self.c.warn(name, "content.good_performance", "列表为空（该场景无优秀表现，可能正常）")
                else:
                    self.c.fail(name, "content.good_performance", f"不是list，实际类型: {type(good).__name__}")

            if need is not None:
                if isinstance(need, list):
                    cnt = len(need)
                    if cnt > 0:
                        self.c.ok(name, f"content.need_strengthen (共{cnt}条)")
                    else:
                        self.c.warn(name, "content.need_strengthen", "列表为空（该场景无需加强项，可能正常）")
                else:
                    self.c.fail(name, "content.need_strengthen", f"不是list，实际类型: {type(need).__name__}")

        # good_default 和 need_default（默认文案不能为空）
        good_default = content.get("good_default")
        if isinstance(good_default, list) and len(good_default) > 0:
            self.c.ok(name, "content.good_default非空")
        else:
            self.c.fail(name, "content.good_default", f"值为: {good_default}")

        need_default = content.get("need_default")
        if isinstance(need_default, list) and len(need_default) > 0:
            self.c.ok(name, "content.need_default非空")
        else:
            self.c.fail(name, "content.need_default", f"值为: {need_default}")

        # watermark字段应为null（需求中未提到，默认不展示）
        if content.get("watermark") is not None:
            self.c.warn(name, "content.watermark", f"watermark不为null: {content.get('watermark')}")
        else:
            self.c.ok(name, "content.watermark==null")


# ==========================================
# TC-05: PRE_POST_SETTLE
# ==========================================

class PrePostSettleValidator:
    """
    需求: 结尾帧，data必须为null
    """

    def __init__(self, checker: Checker):
        self.c = checker

    def validate(self, template: dict):
        name = "PRE_POST_SETTLE"
        self.c.check_null(template.get("data"), name, "data==null")


# ==========================================
# TC-07: 掌握度与内容对应校验
# ==========================================

class MasteryContentValidator:
    """
    根据文件名中的掌握度标识，校验SUMMARY_PERSON中good/need内容是否符合逻辑
    - excellent_mastery: good_performance应有数据
    - not_mastered: need_strengthen应有数据
    - moderate_mastery: 两者可能都有
    """

    def __init__(self, checker: Checker):
        self.c = checker

    def validate(self, templates: list, mastery_level: str):
        name = "MASTERY_CONTENT"

        person_templates = [
            t for t in templates
            if t.get("template_code") == "SUMMARY_PERSON"
        ]

        if not person_templates:
            # 某些课型没有SUMMARY_PERSON（如100阅读），跳过
            return

        content = (person_templates[0].get("data") or {}).get("content") or {}
        good = content.get("good_performance")
        need = content.get("need_strengthen")

        if mastery_level == "excellent":
            # excellent时good_performance应有数据（可能自拼课全是need，但至少应存在good列表）
            if good and isinstance(good, list) and len(good) > 0:
                self.c.ok(name, "excellent_mastery => good_performance有数据")
            elif isinstance(good, list) and len(good) == 0:
                self.c.warn(
                    name, "excellent_mastery => good_performance",
                    "excellent场景good_performance为空列表，请确认是否符合业务预期"
                )
            else:
                self.c.fail(name, "excellent_mastery => good_performance", "应有数据但为空或null")

        elif mastery_level == "not_mastered":
            if need and isinstance(need, list) and len(need) > 0:
                self.c.ok(name, "not_mastered => need_strengthen有数据")
            else:
                self.c.fail(name, "not_mastered => need_strengthen", "应有数据但为空")

        elif mastery_level == "moderate":
            self.c.warn(
                name,
                "moderate_mastery",
                f"good={len(good) if good else 0}条, need={len(need) if need else 0}条"
            )


# ==========================================
# validator分发映射
# ==========================================

VALIDATOR_MAP = {
    "SUMMARY_PERFECT": SummaryPerfectValidator,
    "SUMMARY_V1": SummaryV1Validator,
    "SUMMARY_V2": SummaryV2Validator,
    "SUMMARY_PERSON": SummaryPersonValidator,
    "PRE_POST_SETTLE": PrePostSettleValidator,
}

IGNORED_CODES = {
    "LISTEN_CHOOSE_PIC",
    "LISTEN_CHOOSE_TEXT",
    "LISTEN_TEXT_CHOOSE_PIC",
    "LISTEN_JUDGE",
    "MATCH",
    "PIC_TEXT_CHOOSE_AUDIO",
    "PIC_TEXT_LISTEN_REPEAT",
    "PIC_TEXT_SPEAK",
}


# ==========================================
# 掌握度推断（从文件名）
# ==========================================

def infer_mastery(filename: str) -> str:
    name = os.path.basename(filename).lower()
    if "excellent" in name:
        return "excellent"
    if "not_mastered" in name or "not mastered" in name:
        return "not_mastered"
    if "moderate" in name:
        return "moderate"
    return "unknown"


# ==========================================
# 单文件校验入口
# ==========================================

def validate_file(filepath: str, report: ExcelReport, mastery_override: str = None):
    print(f"\n{'=' * 60}")
    print(f"文件: {os.path.basename(filepath)}")
    print("=" * 60)

    try:
        with open(filepath, encoding="utf-8") as f:
            data = json.load(f)
    except json.JSONDecodeError as e:
        print(f"× JSON解析失败: {e}")
        report.add(os.path.basename(filepath), "FILE", "JSON解析", "FAIL", str(e))
        return
    except Exception as e:
        print(f"× 文件读取失败: {e}")
        report.add(os.path.basename(filepath), "FILE", "文件读取", "FAIL", str(e))
        return

    checker = Checker(report, filepath)
    mastery = mastery_override or infer_mastery(filepath)

    if mastery != "unknown":
        print(f"掌握度: {mastery}")

    # TC-06: 整体结构
    struct_validator = ResponseStructureValidator(checker)
    if not struct_validator.validate(data):
        return

    templates = data.get("res", [])

    # TC-01~05: 逐个template校验
    for template in templates:
        code = template.get("template_code")

        if code in IGNORED_CODES:
            continue

        validator_cls = VALIDATOR_MAP.get(code)
        if validator_cls:
            validator_cls(checker).validate(template)
        else:
            print(f"  ? [{code}] 未知模板，跳过")

    # TC-07: 掌握度内容对应
    mastery_validator = MasteryContentValidator(checker)
    mastery_validator.validate(templates, mastery)


# ==========================================
# 主函数
# ==========================================

def main():
    parser = argparse.ArgumentParser(
        description="701课后模版接口JSON测试工具"
    )
    parser.add_argument(
        "target",
        help="JSON文件路径或包含JSON文件的目录"
    )
    parser.add_argument(
        "--mastery",
        choices=["excellent", "moderate", "not_mastered"],
        default=None,
        help="手动指定掌握度（不指定则从文件名推断）"
    )
    parser.add_argument(
        "--output",
        default=".",
        help="报告输出目录（默认当前目录）"
    )

    args = parser.parse_args()

    report = ExcelReport()

    target = args.target

    # 收集所有待测文件
    files = []
    if os.path.isdir(target):
        files = sorted(glob.glob(os.path.join(target, "*.json")))
        if not files:
            print(f"目录 {target} 下未找到json文件")
            sys.exit(1)
    elif os.path.isfile(target):
        files = [target]
    else:
        # 尝试glob模式
        files = sorted(glob.glob(target))
        if not files:
            print(f"未找到匹配的文件: {target}")
            sys.exit(1)

    print(f"共找到 {len(files)} 个JSON文件")

    for filepath in files:
        validate_file(filepath, report, mastery_override=args.mastery)

    # 保存报告
    report_path = report.save(args.output)

    print(f"\n{'=' * 60}")
    print(f"测试完成")
    print(f"  PASS : {report.pass_count}")
    print(f"  FAIL : {report.fail_count}")
    print(f"  WARN : {report.warn_count}")
    print(f"  报告 : {report_path}")
    print("=" * 60)

    if report.fail_count > 0:
        sys.exit(1)

    sys.exit(0)


if __name__ == "__main__":
    main()
