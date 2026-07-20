import json
import sys
from pathlib import Path
from openpyxl import Workbook
from datetime import datetime
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

def _is_summary(template_code: str) -> bool:
    return isinstance(template_code, str) and "SUMMARY" in template_code

fills = [
    PatternFill(fill_type="solid", fgColor="DDEBF7"),
    PatternFill(fill_type="solid", fgColor="E2F0D9"),
    PatternFill(fill_type="solid", fgColor="FFF2CC"),
    PatternFill(fill_type="solid", fgColor="FCE4D6"),
    PatternFill(fill_type="solid", fgColor="E4DFEC"),
    PatternFill(fill_type="solid", fgColor="D9EAD3"),
    PatternFill(fill_type="solid", fgColor="F4CCCC"),
    PatternFill(fill_type="solid", fgColor="D0E0E3"),
]

BASE_DIR = Path(__file__).parent
INPUT_DIR = BASE_DIR / "6种课型的接口返回json"

        
def run(suffix=None):
    """
    suffix:
        None -> 处理所有json
        moderate_mastery -> 只处理 *moderate_mastery.json
    """
   

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    if suffix:
        json_files = sorted(INPUT_DIR.glob(f"*{suffix}.json"))
        output_file = BASE_DIR / f"{suffix}_sunchuan_result_{timestamp}.xlsx"
    else:
        json_files = sorted(INPUT_DIR.glob("*.json"))
        output_file = BASE_DIR / f"sunchuan_result_{timestamp}.xlsx"

    if not json_files:
        print(f"未找到符合条件的json文件 suffix={suffix}")
        return

    # 第一遍扫描：收集所有 ability_level 的 key，保持出现顺序
    ability_keys_ordered = []
    ability_keys_seen = set()

    def _collect_ability_keys(node):
        ability = None
        if isinstance(node, dict):
            ability = node.get("ability_level") or {}
        if ability:
            for k in ability:
                if k not in ability_keys_seen:
                    ability_keys_seen.add(k)
                    ability_keys_ordered.append(k)

    for json_file in json_files:
        with open(json_file, "r", encoding="utf-8") as f:
            data = json.load(f)
        for item in data.get("res", []):
            if not _is_summary(item.get("template_code")):
                continue
            knowledges = (item.get("data") or {}).get("knowledges") or {}
            for k, v in knowledges.items():
                if k == "video":
                    continue
                if isinstance(v, list):
                    for x in v:
                        _collect_ability_keys(x)
                elif isinstance(v, dict):
                    _collect_ability_keys(v)

    wb = Workbook()
    ws = wb.active
    ws.title = "result"

    BASE_HEADERS = ["文件名", "template_code", "knowledge_key", "knowledge_name", "knowledge_id", "knowledge_type", "video", "audio"]
    ws.append(BASE_HEADERS + ability_keys_ordered)

    # 第二遍：写数据
    for json_file in json_files:

        print(f"处理：{json_file.name}")

        with open(json_file, "r", encoding="utf-8") as f:
            data = json.load(f)

        for item in data.get("res", []):

            if not _is_summary(item.get("template_code")):
                continue

            template_code = item.get("template_code")

            data_obj = item.get("data") or {}
            knowledges = data_obj.get("knowledges") or {}
            content = data_obj.get("content") or {}

            # 从 knowledges.video 或 content.video 提取 video
            video = knowledges.get("video") or content.get("video") or ""
            # 从 content.audio 提取 audio
            audio = content.get("audio") or ""

            # 如果 knowledges 除 video 外没有其他知识点，只写一行 video/audio
            non_video_keys = [k for k in knowledges if k != "video"]

            if not non_video_keys:
                ws.append([
                    json_file.stem, template_code, "", "", "", "",
                    video, audio,
                ] + [""] * len(ability_keys_ordered))
                continue

            for knowledge_key, value in knowledges.items():

                if knowledge_key == "video":
                    continue

                if not value:
                    continue

                if isinstance(value, list):

                    for x in value:
                        ability = x.get("ability_level") or {}
                        ws.append([
                            json_file.stem, template_code, knowledge_key,
                            x.get("knowledge_name"), x.get("knowledge_id"), x.get("knowledge_type"),
                            video, audio,
                        ] + [ability.get(k) for k in ability_keys_ordered])

                elif isinstance(value, dict):

                    ability = value.get("ability_level") or {}
                    ws.append([
                        json_file.stem, template_code, knowledge_key,
                        value.get("knowledge_name"), value.get("knowledge_id"), value.get("knowledge_type"),
                        video, audio,
                    ] + [ability.get(k) for k in ability_keys_ordered])

    # 自动列宽
    for column_cells in ws.columns:

        max_length = 0
        column = column_cells[0].column

        for cell in column_cells:

            if cell.value is not None:
                value = str(cell.value)
                length = sum(2 if ord(ch) > 127 else 1 for ch in value)
                max_length = max(max_length, length)

        ws.column_dimensions[get_column_letter(column)].width = min(max_length + 2, 60)

    # 相同文件名使用同一种颜色
    file_fill_map = {}
    fill_index = 0

    for row in ws.iter_rows(min_row=2):

        file_name = row[0].value

        if file_name not in file_fill_map:
            file_fill_map[file_name] = fills[fill_index % len(fills)]
            fill_index += 1

        fill = file_fill_map[file_name]

        for cell in row:
            cell.fill = fill

    wb.save(output_file)

    print("=" * 60)
    print(f"处理完成，共处理 {len(json_files)} 个文件")
    print(f"Excel已生成：{output_file}")
    print("=" * 60)


if __name__ == "__main__":

    suffix = sys.argv[1] if len(sys.argv) > 1 else None

    run(suffix)