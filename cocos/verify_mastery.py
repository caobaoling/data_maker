"""
verify_mastery.py
作者: bao0
描述: 验证课后推荐题接口返回数据是否符合掌握度起点规则

需求规则：
  同一知识类型+能力标签下，掌握度起点相同的随机取一个题型分类
  - 掌握程度=1: 返回起点含1及1以上的题型
  - 掌握程度=2: 返回起点含2及2以上的题型
  - 掌握程度=3: 返回起点含3及3以上的题型
  - 掌握程度=4: 接口不返回该知识点

数据来源：
  - 源数据: DB midplatform_user_learning.user_appoint_knowledge_mastery
  - 接口结果: json.json
  - 题型映射: 题型定义映射关系.xlsx  (题型编码 → 中文名)
  - 对应规则: 需求的对应关系.xlsx    (知识类型+能力标签+掌握度起点 → 允许题型列表)
"""

import json
import os
import sys
import pymysql

sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook
from collections import defaultdict

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
from common.db_utils import load_config

# ─────────────────────────────────────────────────────────
# 配置
# ─────────────────────────────────────────────────────────
APPOINT_ID   = '100'
LESSON_TYPE  = 'reading_class'  # 课型-en，留空 '' 则不过滤（可选值见 GC_L0~L2_有效学习目标.xlsx 课型-en 列）
LESSON_LEVEL = 'Level 2'        # Level，留空 '' 则不过滤

# DB mastery 字段值 → 数字
MASTERY_MAP = {
    'not_mastered':     1,
    'moderate_mastery': 2,
    'excellent_mastery':3,
    'perfect_mastery':  4,
}

# DB knowledge_type 英文 → 需求xlsx 中文

KNOWLEDGE_TYPE_MAP = {
    'word_sense': '词汇',
    'phrase_sense': '词汇',
    'sentence': '句型',
    'sentence_pattern_structure': '句型',
    'alphabet_rule': '字母',
    'alphabet_rule_example_word': '字母代表词',
    'phonics_rule': '自然拼读',
    'phonics_rule_example_word': '自拼代表词',
    'chant': '歌谣',
    'conversation': '对话',
    'conversation_sentence': '对话某轮中的某一句',
    'functional_language': '对话功能句',
    'reading': '阅读',
    'reading_content': '阅读每一页的内容',
    'sentence_pattern_example_sentence': '句式例句'
}

# DB ability_type 英文 → 需求xlsx 中文
ABILITY_TYPE_MAP = {
    'INPUT-LISTEN': '输入-听',
    'OUTPUT-SPEAK': '输出-说',
    'LETTER-MATCH': '字母-大小写匹配',
    'LETTER-NAME': '字母-命名',
    'PHONICS-READ': '自拼-认读',
    'PHONICS-WRITE': '自拼-拼写',
    'PHONICS-SPELLING': '自拼-拼读',
    'INPUT-READ': '输入-读',
    'INTERACTION-SPEAK': '交际-说'
}

# ─────────────────────────────────────────────────────────
# 1. 读取题型编码 → 中文名称映射
#    来源: 题型定义映射关系.xlsx  列: [题型编码, 题型]
# ─────────────────────────────────────────────────────────
def load_question_type_map() -> tuple:
    """
    返回:
      mapping:         {题型编码: 中文名称}，如 {'TS_1001': '听音选图'}
      reverse_mapping: {中文名称: 题型编码}，如 {'听音选图': 'TS_1001'}
    """
    path = os.path.join(os.path.dirname(__file__), '题型定义映射关系.xlsx')
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    mapping = {}
    reverse_mapping = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:  # 跳过表头
            continue
        code, name = row[0], row[1]
        if code and name:
            mapping[str(code).strip()] = str(name).strip()
            reverse_mapping[str(name).strip()] = str(code).strip()
    wb.close()
    return mapping, reverse_mapping


# 需求xlsx题型分类名称 → 映射xlsx标准名称
# 用于处理需求文档中的别名（如"听音选图2选1"/"听音选图3选1" 均对应 "听音选图"）
QTYPE_NAME_NORMALIZE = {
    '听音选图2选1':      '听音选图',
    '听音选图3选1':      '听音选图',
    '听音选图3选1':      '听音选图',
    '听音选文2选1':      '听音选文',    
    '听音选文3选1':      '听音选文',
    '听音选图文2选1':      '听音选图文',
    '听文选图2选1':      '听文选图',
    '听音选图文3选1':      '听音选图文',
    '图文选音2选1':      '图文选音',
    '听文选文【细节题】2选1':      '听文选文',
    '听文选图【细节题】2选1':      '听文选图',
    '听文选文【细节题】3选1':      '听文选文',
    '听文选图【细节题】3选1':      '听文选图',
    '看图选音2选1':      '看图选音',
    '看图选音3选1':      '看图选音',
    '图音连线3-3':      '图音连线',
    '图图连线3-3':      '图图连线',
    '听图判断':      '听图判断',
    '听图跟读':        '听图跟读',
    '看图开口':          '看图开口',
}


# ─────────────────────────────────────────────────────────
# 2. 读取需求对应关系
#    来源: 需求的对应关系.xlsx
#    列: 知识类型 | 能力标签 | 掌握度起点 | 题型分类
#    返回: {(知识类型中文, 能力标签中文, 掌握度起点数字): [标准题型中文名, ...]}
# ─────────────────────────────────────────────────────────
def load_requirement_rules() -> tuple:
    """
    解析需求xlsx，处理合并单元格（None继承上一行值）。
    掌握度起点 '1, 2' 拆分为 [1, 2]，数字直接转为 [n]。
    '--' 表示该起点无对应题型（掌握程度=1时输出-说不返回题）。
    题型名称通过 QTYPE_NAME_NORMALIZE 标准化为映射xlsx中的名称。

    返回:
      rules:        {(知识类型, 能力标签, 起点数字): [标准题型名, ...]}
      level_groups: {(知识类型, 能力标签): [(起点数字列表, has_qtypes), ...]}
                    每个元素对应xlsx中的一行，has_qtypes=True表示该层级有题型（非'--'）
                    用于准确计算"期望题数"（'1,2'算同一层级，'--'层不计入题数）
    """
    path = os.path.join(os.path.dirname(__file__), '需求的对应关系.xlsx')
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    rules = defaultdict(list)
    # {(kt, al): [(mastery_points_list, has_qtypes), ...]}，每行追加一个元素
    level_groups = defaultdict(list)
    last_knowledge_type = None
    last_ability_label  = None

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:  # 跳过表头
            continue

        kt_raw, al_raw, mastery_raw, qtype_name = row[0], row[1], row[2], row[3]

        # 处理合并单元格：None 继承上一行
        if kt_raw is not None:
            last_knowledge_type = str(kt_raw).strip()
        if al_raw is not None:
            last_ability_label = str(al_raw).strip()

        kt = last_knowledge_type
        al = last_ability_label

        if mastery_raw is None or qtype_name is None:
            continue

        # 解析掌握度起点，支持 '1, 2' 或单个数字
        mastery_str = str(mastery_raw).strip()
        try:
            mastery_points = [int(x.strip()) for x in mastery_str.split(',')]
        except ValueError:
            continue

        qtype_name = str(qtype_name).strip()
        if qtype_name == '--':
            for mp in mastery_points:
                # 记录为空列表，表示该起点不应返回任何题型
                key = (kt, al, mp)
                if key not in rules:
                    rules[key] = []
            # 该行是'--'层级，has_qtypes=False
            level_groups[(kt, al)].append((mastery_points, False))
            continue

        # 标准化题型名称
        std_name = QTYPE_NAME_NORMALIZE.get(qtype_name, qtype_name)

        for mp in mastery_points:
            key = (kt, al, mp)
            if std_name not in rules[key]:
                rules[key].append(std_name)

        # 该行有题型，has_qtypes=True；同一掌握度起点组（如'1,2'）只记录一次层级
        group_key = (kt, al)
        existing_points = [pts for pts, _ in level_groups[group_key]]
        if mastery_points not in existing_points:
            level_groups[group_key].append((mastery_points, True))

    wb.close()
    return dict(rules), dict(level_groups)


# ─────────────────────────────────────────────────────────
# 3. 从数据库读取掌握度数据
# ─────────────────────────────────────────────────────────
def load_db_mastery(appoint_id: str) -> list:
    """
    查询 user_appoint_knowledge_mastery，返回列表：
    [{knowledge_id, knowledge_type, ability_type, mastery(数字1-4)}]
    """
    cfg = load_config()
    conn = pymysql.connect(
        host=cfg['host'],
        port=cfg['port'],
        user=cfg['user'],
        password=cfg['password'],
        database='midplatform_user_learning',
        charset='utf8mb4',
        cursorclass=pymysql.cursors.DictCursor
    )
    sql = """
        SELECT *
        FROM `user_appoint_knowledge_mastery`
        WHERE `appoint_id` = %s
        ORDER BY `knowledge_type` DESC
        LIMIT 1000
    """
    with conn:
        with conn.cursor() as cur:
            cur.execute(sql, (appoint_id,))
            rows = cur.fetchall()

    result = []
    for row in rows:
        mastery_str = row.get('mastery', '')
        mastery_num = MASTERY_MAP.get(mastery_str)
        if mastery_num is None:
            print(f"  [警告] 未知mastery值: {mastery_str}，跳过 knowledge_id={row.get('knowledge_id')}")
            continue
        result.append({
            'knowledge_id':   str(row['knowledge_id']),
            'knowledge_type': row['knowledge_type'],   # 英文，如 word_sense
            'ability_type':   row['ability_type'],     # 英文，如 OUTPUT-SPEAK
            'mastery':        mastery_num,
        })
    return result


# ─────────────────────────────────────────────────────────
# 4. 解析接口返回数据 json.json
# ─────────────────────────────────────────────────────────
def load_api_questions(json_file: str = 'json.json') -> list:
    """
    解析 json.json，提取每道题：
    {template_code, question_type(编码), knowledge_type(英文), ability_type(英文), knowledge_ids}
    忽略 SUMMARY_V1非完美总结页/ SUMMARY_V2非完美总结页/ SUMMARY_PERSON个性化总结页/ PRE_POST_SETTLE结算页/SUMMARY_PERFECT完美总结页 等非题目模板。
    """
    json_path = os.path.join(os.path.dirname(__file__), json_file)
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    skip_templates = {'SUMMARY_V1','SUMMARY_V2', 'SUMMARY_PERSON', 'PRE_POST_SETTLE','SUMMARY_PERFECT'}
    questions = []

    for item in data.get('res', []):
        template_code     = item.get('template_code', '')
        template_kt       = item.get('knowledge_type', '')  # template 级别的 knowledge_type
        if template_code in skip_templates:
            continue
        template_data = item.get('data')
        if not template_data:
            continue

        for q in template_data.get('questions', []):
            question_type = q.get('question_type', '')
            tag_list      = q.get('tag_list', [])
            point_list    = q.get('point_list', [])

            # 取 ability_type：tag_list 中 category_code=ability 的 code
            ability_type = ''
            for tag in tag_list:
                if tag.get('category_code') == 'ability':
                    ability_type = tag.get('code', '')
                    break

            # 用 template 级别的 knowledge_type 过滤 point_list，
            # 只取 point_type_name_en == template_kt 的 point_id
            if template_kt:
                knowledge_ids = [
                    str(p.get('point_id', ''))
                    for p in point_list
                    if p.get('point_type_name_en') == template_kt
                ]
                knowledge_type = template_kt
            else:
                # template 无 knowledge_type 时兜底取全部
                knowledge_ids  = [str(p.get('point_id', '')) for p in point_list]
                knowledge_type = point_list[0].get('point_type_name_en', '') if point_list else ''

            questions.append({
                'template_code':  template_code,
                'question_type':  question_type,
                'knowledge_type': knowledge_type,
                'ability_type':   ability_type,
                'knowledge_ids':  knowledge_ids,
            })

    return questions


# ─────────────────────────────────────────────────────────
# 5. 读取本节课知识点范围（根据课型-en + Level）
# ─────────────────────────────────────────────────────────
def load_lesson_scope(lesson_type_en: str, level: str) -> set:
    """
    从 GC_L0~L2_有效学习目标.xlsx 读取指定课型-en + Level 的知识点类型和能力项组合。
    列顺序: 课型 | 课型-en | Level | point_type_name_en | ability_code
    匹配字段: 课型-en（如 reading_class）+ Level（如 Level 2）
    返回: {(point_type_name_en, ability_code_normalized), ...}
    ability_code 统一将下划线替换为连字符（如 INPUT_LISTEN → INPUT-LISTEN）
    """
    path = os.path.join(os.path.dirname(__file__), 'GC_L0~L2_有效学习目标.xlsx')
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    result = set()
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        kt_en_raw, lv_raw, pt_en, ab_code = row[1], row[2], row[3], row[4]
        if str(kt_en_raw).strip().lower() != lesson_type_en.strip().lower() or str(lv_raw).strip().lower() != level.strip().lower():
            continue
        if pt_en and ab_code:
            ab_normalized = str(ab_code).strip().replace('_', '-')
            result.add((str(pt_en).strip(), ab_normalized))
    wb.close()
    return result


# ─────────────────────────────────────────────────────────
# 6. 核心校验
# ─────────────────────────────────────────────────────────

def _calc_expected_lianxian(n: int) -> int:
    """
    mastery=3 层，知识点数 N > 1 时，连线题期望道数：
      N in [2, 4]  → 1道
      N in [5, 8]  → 2道
      N in [9, 12] → 3道
      其余(N>12)   → 0道（取不到，视为空）
    """
    if 2 <= n <= 4:
        return 1
    if 5 <= n <= 8:
        return 2
    if 9 <= n <= 12:
        return 3
    return 0


# 连线类题型的标准中文名集合（这类题型可将多个知识点融合为一道题）
LIANXIAN_QTYPE_NAMES = {'图音连线', '图图连线','图图连线3-3', '图音连线3-3'}


def _get_used_qt_names(lvl_results: dict) -> set:
    """返回已在低层级用过的题型中文名集合，用于高层级过滤。"""
    used = set()
    for qt_name_set in lvl_results.values():
        used.update(qt_name_set)
    return used


def verify(db_rows: list, api_questions: list, qt_map: dict, qt_reverse_map: dict, rules: dict, level_groups: dict, lesson_scope: set = None):
    """
    校验逻辑（按 (knowledge_type, ability_type) 分组整体校验）：

    知识点个数=1：直接按需求对应关系，每个有效层级出1道题。

    知识点个数>1，按层从低到高（mastery_min → 3）逐层校验：
      mastery=1 层：所有 mastery≤1 的知识点，随机1种题型，每人各1题（取不到为空）
      mastery=2 层：所有 mastery≤2 的知识点，过滤已用题型后随机1种，每人各1题（取不到为空）
      mastery=3 层：所有知识点 N 个，过滤已用题型后：
        N=1       → 随机1种题型取1题
        N in 2~4  → 1道连线题
        N in 5~8  → 2道连线题
        N in 9~12 → 3道连线题
        取不到为空（不报错）

    mastery=4：接口不应返回该组任何题目
    """

    # ── Step1：按中文 (kt_cn, at_cn) 分组，合并同类知识类型（如 word_sense+phrase_sense→词汇）──
    # group_db: {(kt_cn, at_cn): [{'knowledge_id', 'mastery', 'kt_en', 'at_en'}, ...]}
    group_db = defaultdict(list)
    for row in db_rows:
        kt_cn = KNOWLEDGE_TYPE_MAP.get(row['knowledge_type'], row['knowledge_type'])
        at_cn = ABILITY_TYPE_MAP.get(row['ability_type'], row['ability_type'])
        group_db[(kt_cn, at_cn)].append({
            'knowledge_id': row['knowledge_id'],
            'mastery':      row['mastery'],
            'kt_en':        row['knowledge_type'],
            'at_en':        row['ability_type'],
        })

    skipped_api_groups = []  # lesson_scope 过滤时被跳过的接口分组

    # ── Step2：构建接口数据索引，同样按英文 (kt_en, at_en) 分组 ──
    # api_by_group_en: {(kt_en, at_en): [{'question_type', 'knowledge_ids'}, ...]}
    # 同时保留中文 key 索引供后续校验使用
    api_by_group = defaultdict(list)
    api_by_group_en = defaultdict(list)
    for q in api_questions:
        kt_cn = KNOWLEDGE_TYPE_MAP.get(q['knowledge_type'], q['knowledge_type'])
        at_cn = ABILITY_TYPE_MAP.get(q['ability_type'], q['ability_type'])
        api_by_group[(kt_cn, at_cn)].append({
            'question_type': q['question_type'],
            'knowledge_ids': q['knowledge_ids'],
        })
        api_by_group_en[(q['knowledge_type'], q['ability_type'])].append({
            'question_type': q['question_type'],
            'knowledge_ids': q['knowledge_ids'],
        })

    # ── 若指定了课程范围（lesson_scope），双向过滤：
    #    1. group_db：只保留该课型涉及的知识点组合（DB 侧过滤）
    #    2. api_by_group：只保留 lesson_scope 中有的 (kt_en, at_en) 对应题目（接口侧过滤）
    #       接口返回但 xlsx 没有的知识点+能力项 → 跳过，不参与校验
    if lesson_scope:
        # 过滤 DB 分组
        group_db = {
            (kt_cn, at_cn): members
            for (kt_cn, at_cn), members in group_db.items()
            if any((m['kt_en'], m['at_en']) in lesson_scope for m in members)
        }
        # 过滤接口分组：重建 api_by_group，只保留 lesson_scope 内的 (kt_en, at_en)
        api_by_group = defaultdict(list)
        skipped_api_groups = []
        for (kt_en, at_en), qs in api_by_group_en.items():
            if (kt_en, at_en) in lesson_scope:
                kt_cn = KNOWLEDGE_TYPE_MAP.get(kt_en, kt_en)
                at_cn = ABILITY_TYPE_MAP.get(at_en, at_en)
                api_by_group[(kt_cn, at_cn)].extend(qs)
            else:
                skipped_api_groups.append((kt_en, at_en, len(qs)))

    print("\n" + "=" * 72)
    print("  掌握度起点校验报告（按知识点分组）")
    print("=" * 72)

    if lesson_scope:
        print(f"\n  [本节课范围] 共 {len(lesson_scope)} 组 知识点类型+能力项：")
        for pt, ab in sorted(lesson_scope):
            print(f"    point_type={pt} | ability={ab}")
        if skipped_api_groups:
            print(f"\n  [跳过] 接口返回但不在本节课范围内，共 {len(skipped_api_groups)} 组（不参与校验）：")
            for kt_en, at_en, cnt in sorted(skipped_api_groups):
                print(f"    point_type={kt_en} | ability={at_en} | 题数={cnt}")

    # 辅助：将题型中文名列表转为 [编码(中文名), ...] 格式
    def _fmt_qt_names(names):
        return [f"{qt_reverse_map.get(n, '?')}({n})" for n in names]

    pass_count = 0
    fail_count = 0
    skip_count = 0

    for (kt_cn, at_cn), members in sorted(group_db.items()):
        # 分组 key 已是中文，直接使用
        group_key = (kt_cn, at_cn)
        n_points  = len(members)
        api_qs    = api_by_group.get((kt_cn, at_cn), [])

        # 取该组最低 mastery（决定出题层级范围）
        mastery_min = min(m['mastery'] for m in members)

        # 打印时展示原始英文类型（可能合并了多种）
        kt_en_types = sorted({m['kt_en'] for m in members})
        at_en_types = sorted({m['at_en'] for m in members})
        print(f"\n  ── 分组: knowledge_type={kt_cn}({kt_en_types}) | ability={at_cn}({at_en_types})")
        print(f"     知识点个数={n_points} | 最低mastery={mastery_min}")
        for m in members:
            print(f"       knowledge_id={m['knowledge_id']} | {m['kt_en']} | {m['at_en']} | mastery={m['mastery']}")

        # ── mastery=4：接口不应返回任何题目 ──
        if mastery_min == 4:
            if not api_qs:
                pass_count += 1
                print(f"     [通过] mastery=4，接口未返回（符合需求）")
            else:
                fail_count += 1
                names = [f"{q['question_type']}({qt_map.get(q['question_type'], q['question_type'])})" for q in api_qs]
                print(f"     [失败] mastery=4，接口不应返回，但实际返回: {names}")
            continue

        # ── 查该组合是否在需求规则中定义 ──
        rule_defined = any(
            (kt_cn, at_cn, m) in rules
            for m in range(mastery_min, 4)
        )
        if not rule_defined:
            skip_count += 1
            actual_names = [
                f"{q['question_type']}({qt_map.get(q['question_type'], q['question_type'])})"
                for q in api_qs
            ]
            print(f"     [跳过] 需求xlsx未定义该组合 | 实际返回{len(api_qs)}道: {actual_names}")
            continue

        # ════════════════════════════════════════════════
        # 知识点个数 = 1：直接按层级逐层出1道题
        # ════════════════════════════════════════════════
        if n_points == 1:
            kid = members[0]['knowledge_id']
            mastery = members[0]['mastery']

            # 总体校验：题数 + 题型合法性
            # 期望题数：逐层过滤，过滤后无新可用题型则该层不出题（与知识点>1逻辑一致）
            expected_q_count = 0
            _used_accum_1: set = set()
            for lv in range(mastery, 4):
                lv_allowed = rules.get((kt_cn, at_cn, lv), [])
                if not lv_allowed:
                    continue  # '--'，不出题
                lv_avail = [t for t in lv_allowed if t not in _used_accum_1]
                if not lv_avail:
                    continue  # 过滤后无新题型，不出题
                expected_q_count += 1
                _used_accum_1.add(lv_avail[0])
            actual_q_count = len(api_qs)
            all_allowed_names = set()
            for m in range(mastery, 4):
                all_allowed_names.update(rules.get((kt_cn, at_cn, m), []))

            invalid_qs = [
                q for q in api_qs
                if qt_map.get(q['question_type'], q['question_type']) not in all_allowed_names
            ]

            # 构建期望层级说明（显示 编码(中文名)）
            expect_lines_1 = []
            for lv in range(mastery, 4):
                lv_allowed = rules.get((kt_cn, at_cn, lv), [])
                if not lv_allowed:
                    expect_lines_1.append(f"mastery={lv}: --")
                else:
                    expect_lines_1.append(f"mastery={lv}: {_fmt_qt_names(lv_allowed)}")

            # 构建实际返回说明
            actual_lines_1 = [
                f"{q['question_type']}({qt_map.get(q['question_type'], q['question_type'])}) kids={q.get('knowledge_ids', [])}"
                for q in api_qs
            ]

            if not api_qs and expected_q_count == 0:
                pass_count += 1
                print(f"     [通过] 知识点=1 | mastery={mastery} | 期望{expected_q_count}道，实际{actual_q_count}道")
                print(f"            期望层级: {' | '.join(expect_lines_1)}")
                print(f"            规则要求不返回，接口未返回")
            elif not api_qs:
                fail_count += 1
                print(f"     [失败] 知识点=1 | mastery={mastery} | 期望{expected_q_count}道，实际{actual_q_count}道")
                print(f"            期望层级: {' | '.join(expect_lines_1)}")
                print(f"            实际返回: 无")
            elif invalid_qs:
                fail_count += 1
                inv_names = [
                    f"{q['question_type']}({qt_map.get(q['question_type'], q['question_type'])})"
                    for q in invalid_qs
                ]
                print(f"     [失败] 知识点=1 | mastery={mastery} | 期望{expected_q_count}道，实际{actual_q_count}道")
                print(f"            期望层级: {' | '.join(expect_lines_1)}")
                print(f"            含不合规题型: {inv_names}")
                for al in actual_lines_1:
                    print(f"              {al}")
            elif actual_q_count != expected_q_count:
                fail_count += 1
                print(f"     [失败] 知识点=1 | mastery={mastery} | 期望{expected_q_count}道，实际{actual_q_count}道")
                print(f"            期望层级: {' | '.join(expect_lines_1)}")
                for al in actual_lines_1:
                    print(f"              {al}")
            else:
                pass_count += 1
                print(f"     [通过] 知识点=1 | mastery={mastery} | 期望{expected_q_count}道，实际{actual_q_count}道")
                print(f"            期望层级: {' | '.join(expect_lines_1)}")
                for al in actual_lines_1:
                    print(f"              {al}")

        # ════════════════════════════════════════════════
        # 知识点个数 > 1：分层校验
        #
        # 规则（以 mastery=3 是否有连线题为分支）：
        #
        # 分支A：mastery=3 允许题型中包含连线题
        #   → 只出连线题（按知识点数区间计算道数），跳过 mastery=1/2 层
        #
        # 分支B：mastery=3 允许题型中无连线题
        #   → 按 mastery=1 → mastery=2 → mastery=3 顺序出题
        #      每层过滤掉前面所有层已用的题型，同层统一1种题型
        # ════════════════════════════════════════════════
        else:
            group_errors = []
            # 已被归属的题目索引，避免同一道题被多层重复计入
            claimed_indices: set = set()
            # 各层实际使用的题型名，用于逐层过滤
            used_qt_by_lvl: dict = {}

            # ── 预判：mastery=3 层的允许题型中是否含连线题 ──
            allowed_lv3 = rules.get((kt_cn, at_cn, 3), [])
            lx_in_lv3   = [t for t in allowed_lv3 if t in LIANXIAN_QTYPE_NAMES]
            use_lianxian = bool(lx_in_lv3)  # True=走分支A，False=走分支B

            if use_lianxian:
                # ════════════════════════════════
                # 分支A：mastery=3 有连线题
                #   mastery=1/2 层：按分支B方式出非连线题（逐层过滤）
                #   mastery=3 层：出连线题（按知识点数区间）
                # ════════════════════════════════

                # ── 先校验 mastery=1/2 层的非连线题 ──
                for lvl in range(mastery_min, 3):  # 只遍历到 mastery=2
                    rule_key = (kt_cn, at_cn, lvl)
                    if rule_key not in rules:
                        continue
                    allowed = rules[rule_key]
                    lvl_members = [m for m in members if m['mastery'] <= lvl]
                    if not lvl_members or not allowed:
                        used_qt_by_lvl[lvl] = set()
                        continue

                    used_names: set = set()
                    for past_lvl in range(mastery_min, lvl):
                        used_names.update(used_qt_by_lvl.get(past_lvl, set()))
                    available_allowed = [t for t in allowed if t not in used_names]
                    n_lvl = len(lvl_members)

                    pick_from = available_allowed if available_allowed else allowed
                    lvl_api_qs = []
                    for i, q in enumerate(api_qs):
                        if i in claimed_indices:
                            continue
                        if qt_map.get(q['question_type'], q['question_type']) in pick_from:
                            lvl_api_qs.append(q)
                            claimed_indices.add(i)
                            if len(lvl_api_qs) >= n_lvl:
                                break

                    lvl_api_qt_names = {
                        qt_map.get(q['question_type'], q['question_type'])
                        for q in lvl_api_qs
                    }
                    if not lvl_api_qs:
                        used_qt_by_lvl[lvl] = set()
                        continue

                    if len(lvl_api_qt_names) > 1:
                        group_errors.append(
                            f"层级mastery={lvl} | 同组出现多种题型（应统一为1种）: {sorted(lvl_api_qt_names)}"
                        )
                    invalid_names = lvl_api_qt_names - set(allowed)
                    if invalid_names:
                        group_errors.append(
                            f"层级mastery={lvl} | 含不合规题型: {sorted(invalid_names)} | 允许: {allowed}"
                        )
                    reused = lvl_api_qt_names & used_names
                    if reused and available_allowed:
                        group_errors.append(
                            f"层级mastery={lvl} | 使用了低层级已用题型: {sorted(reused)}"
                            f"（已用: {sorted(used_names)}，可用: {available_allowed}）"
                        )
                    if len(lvl_api_qs) != n_lvl:
                        group_errors.append(
                            f"层级mastery={lvl} | 期望{n_lvl}道（每知识点1题），实际{len(lvl_api_qs)}道"
                        )
                    used_qt_by_lvl[lvl] = lvl_api_qt_names

                # ── 再校验 mastery=3 层的连线题 ──
                lv3_members = [m for m in members if m['mastery'] <= 3]
                n_lv3 = len(lv3_members)
                expected_lx = _calc_expected_lianxian(n_lv3)

                lx_api_qs = [
                    q for i, q in enumerate(api_qs)
                    if i not in claimed_indices
                    and qt_map.get(q['question_type'], q['question_type']) in LIANXIAN_QTYPE_NAMES
                ]
                if expected_lx == 0:
                    group_errors.append(
                        f"mastery=3 | N={n_lv3} 超出连线题区间(>12)，不应返回连线题"
                    )
                elif len(lx_api_qs) != expected_lx:
                    group_errors.append(
                        f"mastery=3 | N={n_lv3}，期望连线题{expected_lx}道，实际{len(lx_api_qs)}道"
                    )
                lx_qt_names = {qt_map.get(q['question_type'], q['question_type']) for q in lx_api_qs}
                invalid_lx = lx_qt_names - set(allowed_lv3)
                if invalid_lx:
                    group_errors.append(
                        f"mastery=3 | 连线题含不合规题型: {sorted(invalid_lx)} | 允许: {allowed_lv3}"
                    )

                # 期望总数 = mastery=1/2 各层出题数 + mastery=3 连线题道数
                expected_grand_total = 0
                _used_accum_a: set = set()
                for lv in range(mastery_min, 3):
                    lv_allowed = rules.get((kt_cn, at_cn, lv), [])
                    if not lv_allowed:
                        continue
                    lv_cnt = len([m for m in members if m['mastery'] <= lv])
                    if lv_cnt == 0:
                        continue
                    lv_avail = [t for t in lv_allowed if t not in _used_accum_a]
                    if not lv_avail:
                        continue
                    expected_grand_total += lv_cnt
                    _used_accum_a.add(lv_avail[0])
                expected_grand_total += expected_lx

            else:
                # ════════════════════════════════
                # 分支B：mastery=3 无连线题，按层出题（mastery=1→2→3，逐层过滤）
                # ════════════════════════════════
                for lvl in range(mastery_min, 4):
                    rule_key = (kt_cn, at_cn, lvl)
                    if rule_key not in rules:
                        continue

                    allowed = rules[rule_key]  # 空列表 = '--'

                    # 该层需要出题的知识点（mastery <= lvl）
                    lvl_members = [m for m in members if m['mastery'] <= lvl]
                    if not lvl_members:
                        continue

                    # '--'：不应出题
                    if not allowed:
                        used_qt_by_lvl[lvl] = set()
                        continue

                    # 累计低层已用题型
                    used_names: set = set()
                    for past_lvl in range(mastery_min, lvl):
                        used_names.update(used_qt_by_lvl.get(past_lvl, set()))

                    # 过滤后可用题型（排除低层已用）
                    available_allowed = [t for t in allowed if t not in used_names]

                    n_lvl = len(lvl_members)

                    # 从接口返回中归属本层题目
                    pick_from = available_allowed if available_allowed else allowed
                    lvl_api_qs = []
                    for i, q in enumerate(api_qs):
                        if i in claimed_indices:
                            continue
                        if qt_map.get(q['question_type'], q['question_type']) in pick_from:
                            lvl_api_qs.append(q)
                            claimed_indices.add(i)
                            if len(lvl_api_qs) >= n_lvl:
                                break

                    lvl_api_qt_names = {
                        qt_map.get(q['question_type'], q['question_type'])
                        for q in lvl_api_qs
                    }

                    if not lvl_api_qs:
                        used_qt_by_lvl[lvl] = set()
                        continue

                    # 同层只能出1种题型
                    if len(lvl_api_qt_names) > 1:
                        group_errors.append(
                            f"层级mastery={lvl} | 同组出现多种题型（应统一为1种）: {sorted(lvl_api_qt_names)}"
                        )

                    # 题型合法性
                    invalid_names = lvl_api_qt_names - set(allowed)
                    if invalid_names:
                        group_errors.append(
                            f"层级mastery={lvl} | 含不合规题型: {sorted(invalid_names)} | 允许: {allowed}"
                        )

                    # 不能复用低层已用题型（有其他可选时才报错）
                    reused = lvl_api_qt_names & used_names
                    if reused and available_allowed:
                        group_errors.append(
                            f"层级mastery={lvl} | 使用了低层级已用题型: {sorted(reused)}"
                            f"（已用: {sorted(used_names)}，可用: {available_allowed}）"
                        )

                    # 每个知识点各1道
                    if len(lvl_api_qs) != n_lvl:
                        group_errors.append(
                            f"层级mastery={lvl} | 期望{n_lvl}道（每知识点1题），实际{len(lvl_api_qs)}道"
                        )

                    used_qt_by_lvl[lvl] = lvl_api_qt_names

                # 兜底：接口返回的题目必须全部归属到某个有效层级
                all_valid_qt_names: set = set()
                for lv in range(mastery_min, 4):
                    all_valid_qt_names.update(rules.get((kt_cn, at_cn, lv), []))
                illegal_qs = [
                    f"{q['question_type']}({qt_map.get(q['question_type'], q['question_type'])})"
                    for q in api_qs
                    if qt_map.get(q['question_type'], q['question_type']) not in all_valid_qt_names
                ]
                if illegal_qs:
                    group_errors.append(f"接口返回了不属于任何允许层级的题型: {illegal_qs}")

                # 计算期望总题数（分支B：各层逐层累计）
                expected_grand_total = 0
                _used_accum: set = set()
                for lv in range(mastery_min, 4):
                    lv_allowed = rules.get((kt_cn, at_cn, lv), [])
                    if not lv_allowed:
                        continue
                    lv_cnt = len([m for m in members if m['mastery'] <= lv])
                    if lv_cnt == 0:
                        continue
                    lv_avail = [t for t in lv_allowed if t not in _used_accum]
                    if not lv_avail:
                        continue
                    expected_grand_total += lv_cnt
                    _used_accum.add(lv_avail[0])

            if len(api_qs) != expected_grand_total and not any('整组期望共' in e for e in group_errors):
                group_errors.append(
                    f"整组期望共{expected_grand_total}道题，实际返回{len(api_qs)}道"
                )

            # 构建期望层级说明
            expect_lines = []
            if use_lianxian:
                # 分支A：展示 mastery=1/2 各层（非连线题）+ mastery=3 连线题
                _used_fmt_a: set = set()
                for lv in range(mastery_min, 3):
                    lv_allowed = rules.get((kt_cn, at_cn, lv), [])
                    lv_members_cnt = len([m for m in members if m['mastery'] <= lv])
                    if lv_members_cnt == 0:
                        continue
                    if not lv_allowed:
                        expect_lines.append(f"mastery={lv}: --（不出题）")
                        continue
                    lv_avail = [t for t in lv_allowed if t not in _used_fmt_a]
                    if not lv_avail:
                        expect_lines.append(f"mastery={lv}: 过滤后无新题型，不出题")
                        continue
                    expect_lines.append(
                        f"mastery={lv}: {_fmt_qt_names(lv_avail)} → {lv_members_cnt}道（每知识点1题）"
                    )
                    _used_fmt_a.add(lv_avail[0])
                n_all = len(members)
                expect_lines.append(
                    f"mastery=3(连线题优先): {_fmt_qt_names(lx_in_lv3)} → {_calc_expected_lianxian(n_all)}道"
                )
            else:
                # 分支B：展示各层（mastery=1→2→3）
                for lv in range(mastery_min, 4):
                    lv_allowed = rules.get((kt_cn, at_cn, lv), [])
                    lv_members_cnt = len([m for m in members if m['mastery'] <= lv])
                    if lv_members_cnt == 0:
                        continue
                    if not lv_allowed:
                        expect_lines.append(f"mastery={lv}: --（不出题）")
                        continue
                    expect_lines.append(
                        f"mastery={lv}: {_fmt_qt_names(lv_allowed)} → {lv_members_cnt}道（每知识点1题）"
                    )

            # 构建实际返回说明（每道题显示知识点ID和题型）
            actual_lines = []
            for q in api_qs:
                qt_name = qt_map.get(q['question_type'], q['question_type'])
                kids = q.get('knowledge_ids', [])
                actual_lines.append(f"{q['question_type']}({qt_name}) kids={kids}")

            status = "[失败]" if group_errors else "[通过]"
            print(f"     {status} 知识点={n_points} | 期望{expected_grand_total}道，实际返回{len(api_qs)}道")
            print(f"            期望层级: {' | '.join(expect_lines)}")
            if actual_lines:
                print(f"            实际返回:")
                for al in actual_lines:
                    print(f"              {al}")
            else:
                print(f"            实际返回: 无")
            if group_errors:
                fail_count += 1
                for err in group_errors:
                    print(f"            ✗ {err}")
            else:
                pass_count += 1

    print()
    print("─" * 72)
    print(f"  汇总：通过={pass_count}  失败={fail_count}  跳过={skip_count}")
    print("─" * 72)
    print("=" * 72)


# ─────────────────────────────────────────────────────────
# 入口
# ─────────────────────────────────────────────────────────
if __name__ == '__main__':
    # 支持命令行传参（优先级高于脚本顶部常量）：
    #   python cocos/verify_mastery.py <appoint_id>
    #   python cocos/verify_mastery.py <appoint_id> <课型-en> <Level>
    # 例：python cocos/verify_mastery.py 539673674 reading_class "Level 2"
    appoint_id  = sys.argv[1] if len(sys.argv) > 1 else APPOINT_ID
    lesson_type = sys.argv[2] if len(sys.argv) >= 4 else LESSON_TYPE
    level       = sys.argv[3] if len(sys.argv) >= 4 else LESSON_LEVEL

    print(f"使用 appoint_id={appoint_id}")

    lesson_scope = None
    if lesson_type and level:
        print(f"课程范围: 课型-en={lesson_type} | Level={level}")
        print(f"\n[0] 加载本节课知识点范围 (GC_L0~L2_有效学习目标.xlsx) ...")
        lesson_scope = load_lesson_scope(lesson_type, level)
        if not lesson_scope:
            print(f"    [警告] 未找到 课型-en={lesson_type} Level={level} 的有效学习目标，将校验所有知识点")
        else:
            print(f"    共找到 {len(lesson_scope)} 组 (知识点类型, 能力项) 组合，不在此范围内的知识点将跳过")
            for pair in sorted(lesson_scope):
                print(f"      point_type={pair[0]} | ability={pair[1]}")
    else:
        print(f"未指定课型/Level，校验所有知识点")

    print(f"\n[1] 加载题型编码映射 (题型定义映射关系.xlsx) ...")
    qt_map, qt_reverse_map = load_question_type_map()
    print(f"    共 {len(qt_map)} 条映射")

    print(f"\n[2] 加载需求对应规则 (需求的对应关系.xlsx) ...")
    rules, level_groups = load_requirement_rules()
    print(f"    共 {len(rules)} 条规则组合")
    for k, v in sorted(rules.items()):
        print(f"    {k} → {v}")
    print(f"    层级分组:")
    for k, v in sorted(level_groups.items()):
        print(f"    {k} → {v}")

    print(f"\n[3] 从数据库加载掌握度数据 (appoint_id={appoint_id}) ...")
    db_rows = load_db_mastery(appoint_id)
    print(f"    共 {len(db_rows)} 条记录")
    for row in db_rows:
        print(f"    knowledge_id={row['knowledge_id']} | type={row['knowledge_type']}"
              f" | ability={row['ability_type']} | mastery={row['mastery']}")

    print(f"\n[4] 自动拉取接口数据 ...")
    import importlib.util as _ilu
    _spec = _ilu.spec_from_file_location("fetch_api", os.path.join(os.path.dirname(__file__), "fetch_api.py"))
    _fetch_api = _ilu.module_from_spec(_spec)
    _spec.loader.exec_module(_fetch_api)
    _fetch_api.fetch(appoint_id, appoint_id, '')

    print(f"\n[5] 读取接口返回数据 (json.json) ...")
    api_questions = load_api_questions('json.json')
    print(f"    共 {len(api_questions)} 道题")

    # 打印 mastery=4 的知识点（接口不应返回，提前说明）
    mastery4 = [r for r in db_rows if r['mastery'] == 4]
    if mastery4:
        print(f"\n[说明] 以下知识点 mastery=4(perfect_mastery)，接口不应返回题目：")
        for r in mastery4:
            print(f"       knowledge_id={r['knowledge_id']} | type={r['knowledge_type']} | ability={r['ability_type']}")

    print(f"\n[6] 开始校验 ...")
    verify(db_rows, api_questions, qt_map, qt_reverse_map, rules, level_groups, lesson_scope)
