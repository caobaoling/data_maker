# -*- coding: utf-8 -*-
import sys, os
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook

wb = load_workbook('cocos/需求的对应关系.xlsx', read_only=True, data_only=True)
ws = wb.active
seen = set()
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0: continue
    kt, al = row[0], row[1]
    if kt is not None and (kt, al) not in seen:
        seen.add((kt, al))
        print(f"知识类型: {repr(kt)}  能力标签: {repr(al)}")
wb.close()
