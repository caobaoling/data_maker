import requests
import pandas as pd
import sys
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.formatting.rule import CellIsRule

# -------------------------- 核心配置区 --------------------------
BASE_API = "https://igateway.51suyang.cn/midplatform_textbook/front/lesson/materials/detail"
USER_AGENT = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36"
REFERER = "https://igateway.51suyang.cn/"
TIMEOUT = 8
SIZE_THRESHOLD_MB = 5.0
# 标红样式：红色加粗字体
RED_BOLD_FONT = Font(color="FF0000", bold=True)

# 批量处理的教材lesson_id列表（你指定的6个）
LESSON_ID_LIST = [
    "1918201",
    "1883121",
    "1929081",
    "1919131",
    "1919121",
    "1920141"
]

# 固定公共参数（请替换为浏览器最新抓取的有效token/timestamp）
common_params = {
    "version": "5.0.0.20",
    "phone_type": "pc",
    "user_id": "26750",
    "relId": "26750",
    "teacher_id": "26750",
    "student_id": "44011216",
    "token": "26750_d23ffc9f68cfc93a80d81ec4dab3df8btea",
    "type": "tea",
    "appoint_id": "539694620",
    "lesson_module": "in_class",
    "role": "admin",
    "appkey": "ac_cocos",
    "timestamp": "1783079923891"
}

# 接口名称常量
API_NAME_ORIGIN = "原始接口（无material_status）"
API_NAME_PUBLISH = "发布态接口（material_status=published）"
# ------------------------------------------------------------

def build_two_api_urls(lesson_id: str):
    """根据教材lesson_id拼接两个接口的完整请求地址"""
    origin_params = common_params.copy()
    origin_params["textbook_lesson_id"] = lesson_id
    publish_params = common_params.copy()
    publish_params["textbook_lesson_id"] = lesson_id
    publish_params["material_status"] = "published"

    origin_url = f"{BASE_API}?{'&'.join([f'{k}={v}' for k, v in origin_params.items()])}"
    publish_url = f"{BASE_API}?{'&'.join([f'{k}={v}' for k, v in publish_params.items()])}"
    return origin_url, publish_url

def get_file_size(url: str):
    """请求MP4文件头，获取文件大小，返回格式化结果"""
    headers = {
        "User-Agent": USER_AGENT,
        "Referer": REFERER
    }
    try:
        resp = requests.head(url, headers=headers, timeout=TIMEOUT, allow_redirects=True)
        if resp.status_code == 200 and "Content-Length" in resp.headers:
            byte_size = int(resp.headers["Content-Length"])
            mb_size = round(byte_size / 1024 / 1024, 2)
            return {
                "byte_size": byte_size,
                "mb_size": mb_size,
                "status": "成功"
            }
        else:
            return {
                "byte_size": 0,
                "mb_size": 0,
                "status": f"状态码:{resp.status_code}"
            }
    except Exception as e:
        return {
            "byte_size": 0,
            "mb_size": 0,
            "status": f"异常:{str(e)}"
        }

def traverse_json_extract_mp4(obj, current_page_no=None, result_list=None):
    """递归遍历接口返回的JSON，提取MP4链接+所属页码，自动处理嵌套结构"""
    if result_list is None:
        result_list = []

    if isinstance(obj, dict):
        if "page_no" in obj:
            current_page_no = obj["page_no"]
        for key, value in obj.items():
            if isinstance(value, str) and value.lower().endswith(".mp4") and value.startswith("http"):
                result_list.append({
                    "page_no": current_page_no,
                    "mp4_url": value
                })
            else:
                traverse_json_extract_mp4(value, current_page_no, result_list)
    elif isinstance(obj, list):
        for item in obj:
            traverse_json_extract_mp4(item, current_page_no, result_list)
    return result_list

def process_single_api(api_url: str, api_name: str, lesson_id: str):
    """处理单个接口，返回单文件一行的结构化数据"""
    print(f"  正在处理 {api_name}...")
    try:
        resp = requests.get(api_url, timeout=10)
        resp.raise_for_status()
        json_data = resp.json()

        mp4_with_page = traverse_json_extract_mp4(json_data)
        # 去重（同一个URL+页码只保留一条）
        unique_mp4 = {}
        for item in mp4_with_page:
            key = f"{item['page_no']}_{item['mp4_url']}"
            if key not in unique_mp4:
                unique_mp4[key] = item
        mp4_with_page = list(unique_mp4.values())
        print(f"  {api_name} 提取到 {len(mp4_with_page)} 个唯一MP4文件")

        # 循环获取每个文件的大小
        result_list = []
        for idx, item in enumerate(mp4_with_page, 1):
            url = item["mp4_url"]
            page_no = item["page_no"]
            size_info = get_file_size(url)
            result_list.append({
                "页码": page_no,
                "文件地址": url,
                "文件大小(MB)": size_info["mb_size"],
                "获取状态": size_info["status"]
            })
            print(f"    {idx}. 页码:{page_no} | {url} | {size_info['mb_size']}MB | {size_info['status']}")

        return result_list

    except Exception as e:
        print(f"  ❌ {api_name} 处理失败: {str(e)}")
        return []

def aggregate_by_page(data_list: list, prefix: str):
    """
    按页码分组聚合，同页码的多个文件信息合并到单元格，用换行分隔
    实现：每个页码一行，同页多文件自动换行展示
    """
    if not data_list:
        return pd.DataFrame()

    df = pd.DataFrame(data_list)
    # 按页码分组，聚合所有同页的文件信息
    def agg_func(group):
        return pd.Series({
            f"{prefix}文件地址": "\n".join(group["文件地址"]),
            f"{prefix}文件大小(MB)": "\n".join(group["文件大小(MB)"].astype(str)),
            f"{prefix}获取状态": "\n".join(group["获取状态"]),
            f"{prefix}文件数量": len(group),
            f"{prefix}总大小(MB)": round(group["文件大小(MB)"].sum(), 2),
            f"{prefix}是否有大于5MB文件": "是" if (group["文件大小(MB)"] > SIZE_THRESHOLD_MB).any() else "否"
        })

    df_agg = df.groupby("页码", dropna=False).apply(agg_func).reset_index()
    df_agg = df_agg.rename(columns={"页码": f"{prefix}页码"})
    return df_agg

def merge_origin_publish_by_page(origin_data: list, publish_data: list, lesson_id: str):
    """
    按页码合并原始和发布态数据，同页码的数据放在同一行
    实现：每个页码一行，双接口数据对齐展示
    """
    print(f"  正在按页码合并双接口数据...")

    # 按页码分别聚合两个接口的数据
    df_origin_agg = aggregate_by_page(origin_data, "原始")
    df_publish_agg = aggregate_by_page(publish_data, "发布后")

    # 按页码全外连接，同页码合并到同一行
    df_merged = pd.merge(
        df_origin_agg,
        df_publish_agg,
        left_on="原始页码",
        right_on="发布后页码",
        how="outer"
    )

    # 填充空值为"无"
    df_merged = df_merged.fillna("无")

    # 新增"是否大于5MB"标记：该页码下任意接口有大于5MB的文件则标记为是
    def is_over_5mb(row):
        try:
            origin_has = row["原始是否有大于5MB文件"] == "是"
            publish_has = row["发布后是否有大于5MB文件"] == "是"
            return "是" if origin_has or publish_has else "否"
        except:
            return "否"
    df_merged["是否大于5MB"] = df_merged.apply(is_over_5mb, axis=1)

    # 按页码排序，数值排序避免错乱
    df_merged["sort_page"] = df_merged["原始页码"].where(df_merged["原始页码"] != "无", df_merged["发布后页码"])
    df_merged["sort_page_num"] = pd.to_numeric(df_merged["sort_page"], errors="coerce").fillna(9999)
    df_merged = df_merged.sort_values(by="sort_page_num", ascending=True).reset_index(drop=True)

    # 调整列顺序，符合查看习惯
    df_merged = df_merged[[
        "原始页码",
        "原始文件数量",
        "原始总大小(MB)",
        "原始是否有大于5MB文件",
        "原始文件地址",
        "原始文件大小(MB)",
        "原始获取状态",
        "发布后页码",
        "发布后文件数量",
        "发布后总大小(MB)",
        "发布后是否有大于5MB文件",
        "发布后文件地址",
        "发布后文件大小(MB)",
        "发布后获取状态",
        "是否大于5MB"
    ]]

    # 新增序号列
    df_merged.insert(0, "序号", range(1, len(df_merged)+1))

    print(f"  按页码合并完成，共{len(df_merged)}个页码记录")
    return df_merged

def process_single_lesson(lesson_id: str):
    """处理单个教材（lesson_id）的完整流程：双接口请求→数据获取→按页码合并"""
    print(f"\n" + "="*80)
    print(f"🚀 开始处理教材 lesson_id = {lesson_id}")
    print("="*80)

    # 拼接两个接口URL
    origin_url, publish_url = build_two_api_urls(lesson_id)

    # 处理两个接口
    origin_data = process_single_api(origin_url, API_NAME_ORIGIN, lesson_id)
    publish_data = process_single_api(publish_url, API_NAME_PUBLISH, lesson_id)

    # 按页码合并数据
    merged_df = merge_origin_publish_by_page(origin_data, publish_data, lesson_id)

    # 统计核心指标
    total_pages = len(merged_df)
    over5mb_count = len(merged_df[merged_df["是否大于5MB"] == "是"])
    total_origin_files = merged_df["原始文件数量"].sum() if merged_df["原始文件数量"].dtype != "object" else 0
    total_publish_files = merged_df["发布后文件数量"].sum() if merged_df["发布后文件数量"].dtype != "object" else 0

    print(f"\n✅ 教材 lesson_id={lesson_id} 处理完成")
    print(f"  总页码数: {total_pages} 个")
    print(f"  原始接口总文件数: {total_origin_files} 个")
    print(f"  发布态接口总文件数: {total_publish_files} 个")
    print(f"  大于5MB的页码数: {over5mb_count} 个")

    return {
        "lesson_id": lesson_id,
        "merged_df": merged_df,
        "total_pages": total_pages,
        "over5mb_count": over5mb_count,
        "total_origin_files": total_origin_files,
        "total_publish_files": total_publish_files
    }

def apply_red_formatting(file_path: str):
    """给Excel文件中大于5MB的页码行应用红色加粗标红格式"""
    print(f"\n🎨 正在给大于5MB的页码标红...")
    wb = load_workbook(file_path)

    # 遍历所有工作表，跳过汇总表
    for sheet_name in wb.sheetnames:
        if sheet_name == "全教材汇总":
            continue
        ws = wb[sheet_name]
        print(f"  正在处理工作表: {sheet_name}")

        # 找到总大小列的列号
        origin_total_size_col = None
        publish_total_size_col = None
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col).value
            if cell_value == "原始总大小(MB)":
                origin_total_size_col = col
            elif cell_value == "发布后总大小(MB)":
                publish_total_size_col = col

        # 应用条件格式：总大小大于5的单元格标红加粗
        rule = CellIsRule(
            operator="greaterThan",
            formula=["5"],
            font=RED_BOLD_FONT
        )

        # 给原始总大小列应用格式
        if origin_total_size_col:
            col_letter = ws.cell(row=1, column=origin_total_size_col).column_letter
            ws.conditional_formatting.add(f"{col_letter}2:{col_letter}{ws.max_row}", rule)
        # 给发布后总大小列应用格式
        if publish_total_size_col:
            col_letter = ws.cell(row=1, column=publish_total_size_col).column_letter
            ws.conditional_formatting.add(f"{col_letter}2:{col_letter}{ws.max_row}", rule)

    # 保存修改后的文件
    wb.save(file_path)
    print(f"✅ 标红格式应用完成！")

def export_to_excel(all_lesson_results: list):
    """将所有教材的结果导出到单个Excel，每个教材一个独立工作表"""
    output_file = "教材MP4资源每页一行_标红版.xlsx"
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        # 1. 每个教材的合并数据写入独立工作表
        summary_data = []
        for res in all_lesson_results:
            lid = res["lesson_id"]
            sheet_name = f"教材_{lid}"
            res["merged_df"].to_excel(writer, sheet_name=sheet_name, index=False)
            # 收集汇总数据
            summary_data.append({
                "教材lesson_id": lid,
                "总页码数": res["total_pages"],
                "原始接口总文件数": res["total_origin_files"],
                "发布态接口总文件数": res["total_publish_files"],
                "大于5MB的页码数": res["over5mb_count"]
            })

        # 2. 写入全量汇总工作表
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name="全教材汇总", index=False)

    print(f"\n📊 基础数据导出完成，文件路径：{output_file}")
    return output_file

if __name__ == "__main__":
    print(f"🚀 批量处理启动，共{len(LESSON_ID_LIST)}个教材待处理")
    print(f"待处理教材列表：{LESSON_ID_LIST}")

    all_results = []
    for lesson_id in LESSON_ID_LIST:
        try:
            res = process_single_lesson(lesson_id)
            all_results.append(res)
        except Exception as e:
            print(f"❌ 教材 lesson_id={lesson_id} 处理失败: {str(e)}")
            continue

    # 导出基础Excel
    output_file = export_to_excel(all_results)
    # 应用大于5MB标红格式
    apply_red_formatting(output_file)

    print(f"\n🎉 全部处理完成！最终文件：{output_file}")
    print(f"📌 核心说明：")
    print(f"  1. 【每页一行】每个页码单独占一行，一页有多个文件时，同页的所有文件信息自动换行合并到对应单元格，不会拆分多行")
    print(f"  2. 【双接口对齐】原始接口和发布态接口的同页码数据，强制合并到同一行，所有字段一一对应")
    print(f"  3. 【自动标红】页码总大小大于5MB的单元格自动标红加粗，一眼识别大体积页码")
    print(f"  4. 【完整统计】每个页码新增文件数量、总大小、是否有大于5MB文件的统计字段，无需手动计算")
    print(f"  5. 【批量处理】自动处理6个指定教材，每个教材独立工作表，新增全教材汇总表，方便整体查看")