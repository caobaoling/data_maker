import os
import json
import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side

def extract_json_to_excel():
    # 1. 获取当前时间戳字符串（格式如：20260628_175549）
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # 定义当前目录下的输入文件，以及带时间戳的输出文件名
    input_file = "json.json"
    output_file = f"整理结果_{timestamp}.xlsx"
    
    # 检查当前目录下是否存在 json.json 文件
    if not os.path.exists(input_file):
        print(f"❌ 错误：在当前目录下未找到 【{input_file}】 文件，请检查文件名或路径！")
        return

    print(f"📂 正在读取当前目录下的 {input_file} ...")
    
    # 2. 读取并解析本地 JSON 文件
    with open(input_file, 'r', encoding='utf-8') as f:
        try:
            data = json.load(f)
        except json.JSONDecodeError as e:
            print(f"❌ JSON 格式解析失败，请检查文件内容是否完整。错误信息: {e}")
            return

    rows = []
    
    # 3. 遍历外层的 res 数组
    for item in data.get("res", []):
        template_code = item.get("template_code")
        
        # 过滤：忽略 SUMMARY_V1 和 SUMMARY_PERSON
        if template_code in ["SUMMARY_V1", "SUMMARY_PERSON"]:
            continue
            
        template_data = item.get("data")
        if not template_data:
            continue
            
        questions = template_data.get("questions", [])
        if not questions:
            continue
            
        # 4. 遍历 questions 列表
        for q in questions:
            question_type_name = q.get("question_type_name")
            question_type = q.get("question_type")  # 【新增】：提取题型代号 (如 TSP_1003)
            point_list = q.get("point_list", [])
            tag_list = q.get("tag_list", [])
            
            # 从当前题型的 tag_list 中找出 category_code 为 'ability' 的 code 值
            ability_label = ""
            for tag in tag_list:
                if tag.get("category_code") == "ability":
                    ability_label = tag.get("code")
                    break
            
            # 5. 遍历 point_list（知识点列表），不进行去重
            for point in point_list:
                point_type_name_en = point.get("point_type_name_en")
                point_id = point.get("point_id")
                
                # 组装成结构化的一行数据（加入了题型代码列）
                rows.append({
                    "知识点类型\n(point_type_name_en)": point_type_name_en,
                    "对应能力标签\n(ability)": ability_label,
                    "知识点 ID\n(point_id)": str(point_id),
                    "题型名称\n(question_type_name)": question_type_name,
                    "题型代码\n(question_type)": question_type  # 【新增】
                })
                
    # 6. 转换为 Pandas DataFrame 
    df = pd.DataFrame(rows)
    
    # 7. 分类排序：根据“知识点类型”➔“对应能力标签”➔“知识点 ID”进行多级归类排序
    if not df.empty:
        df = df.sort_values(
            by=[
                "知识点类型\n(point_type_name_en)", 
                "对应能力标签\n(ability)", 
                "知识点 ID\n(point_id)"
            ], 
            ascending=[True, True, True]
        ).reset_index(drop=True)
        
        # 8. 先用 pandas 写入到 Excel 文件中
        df.to_excel(output_file, index=False)
        
        # 9. 使用 openpyxl 进行单元格合并与自适应宽高调整
        wb = load_workbook(output_file)
        ws = wb.active
        
        # 设置全局对齐样式：水平居中、垂直居中、自动换行
        center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # 设置轻微的灰色网格边框
        thin_side = Side(border_style="thin", color="D3D3D3")
        grid_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        # 应用全局对齐和边框样式
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.alignment = center_alignment
                cell.border = grid_border
        
        # 💡 核心逻辑 1：基于指定列的连续相同内容进行垂直合并
        def merge_column_cells(ws, col_index):
            start_row = 2
            max_row = ws.max_row
            
            while start_row <= max_row:
                end_row = start_row
                current_value = ws.cell(row=start_row, column=col_index).value
                
                if current_value is None:
                    start_row += 1
                    continue
                
                while end_row + 1 <= max_row and ws.cell(row=end_row + 1, column=col_index).value == current_value:
                    if col_index > 1:
                        # 确保子层级的合并不会跨越父层级边界
                        parent_style_match = (ws.cell(row=end_row + 1, column=1).value == ws.cell(row=start_row, column=1).value)
                        if not parent_style_match:
                            break
                    end_row += 1
                
                if end_row > start_row:
                    ws.merge_cells(start_row=start_row, end_row=end_row, start_column=col_index, end_column=col_index)
                
                start_row = end_row + 1

        # 分别对前 3 列（知识点类型、对应能力标签、知识点 ID）执行合并
        merge_column_cells(ws, col_index=1)
        merge_column_cells(ws, col_index=2)
        merge_column_cells(ws, col_index=3)
        
        # 💡 核心逻辑 2：智能行高自适应算法
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            max_lines = 1
            for cell in row:
                if cell.value:
                    lines = len(str(cell.value).split('\n'))
                    if lines > max_lines:
                        max_lines = lines
            ws.row_dimensions[row[0].row].height = 18 + (max_lines - 1) * 15

        # 💡 核心逻辑 3：智能列宽自适应算法
        for col in ws.columns:
            col_letter = col[0].column_letter
            max_len = 0
            
            for cell in col:
                if cell.value:
                    val_str = str(cell.value)
                    # 汉字算 2 字节，英文算 1 字节
                    if '\n' in val_str:
                        cell_len = max(sum(2 if ord(char) > 127 else 1 for char in line) for line in val_str.split('\n'))
                    else:
                        cell_len = sum(2 if ord(char) > 127 else 1 for char in val_str)
                    
                    if cell_len > max_len:
                        max_len = cell_len
            
            # 设置自适应宽度，并在最长字符基础上加 3 个字符作为安全缓冲间距
            ws.column_dimensions[col_letter].width = max(min(max_len + 3, 40), 14)
            
        # 10. 保存带有新列、合并、自适应宽高样式的最终 Excel
        wb.save(output_file)
        print(f"🎉 处理完成！成功添加【题型代码 (question_type)】列，且已完成行高列宽自适应调整。")
        print(f"📂 完美的 Excel 表格已安全存至当前目录下：【{output_file}】")
    else:
        print("⚠️ 未提取到符合条件的数据，未生成 Excel 文件。")

# 运行脚本
if __name__ == "__main__":
    extract_json_to_excel()