#!/usr/bin/env python3
# -*- coding:utf-8 -*-

"""
701课后模板自动校验工具

使用方式：

python check_701_template.py 101词汇excellent_mastery.json

输出：

report_xxx.xlsx
"""

import json
import sys
import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font


# ==========================================
# Excel
# ==========================================

class ExcelReport:

    def __init__(self):

        self.wb = Workbook()
        self.ws = self.wb.active

        self.ws.append([
            "Template",
            "Check Item",
            "Result",
            "Message"
        ])

        self.pass_fill = PatternFill(
            fill_type="solid",
            fgColor="C6EFCE"
        )

        self.fail_fill = PatternFill(
            fill_type="solid",
            fgColor="FFC7CE"
        )

        self.pass_count = 0
        self.fail_count = 0

    def add(
            self,
            template,
            item,
            result,
            msg=""
    ):

        self.ws.append([
            template,
            item,
            result,
            msg
        ])

        row = self.ws.max_row

        if result == "PASS":

            self.ws.cell(row, 3).fill = self.pass_fill
            self.pass_count += 1

        else:

            self.ws.cell(row, 3).fill = self.fail_fill
            self.fail_count += 1

    def save(self):

        filename = "report_%s.xlsx" % datetime.now().strftime(
            "%Y%m%d_%H%M%S"
        )

        self.wb.save(filename)

        return filename


# ==========================================
# 公共工具
# ==========================================

class Checker:

    def __init__(self, report):

        self.report = report

    def ok(
            self,
            template,
            item
    ):

        self.report.add(
            template,
            item,
            "PASS",
            ""
        )

        print(
            "√",
            template,
            item
        )

    def fail(
            self,
            template,
            item,
            msg
    ):

        self.report.add(
            template,
            item,
            "FAIL",
            msg
        )

        print(
            "×",
            template,
            item,
            msg
        )

    def check_exist(
            self,
            value,
            template,
            item
    ):

        if value is None:

            self.fail(
                template,
                item,
                "为空"
            )

            return False

        if value == "":

            self.fail(
                template,
                item,
                "为空字符串"
            )

            return False

        self.ok(
            template,
            item
        )

        return True

    def check_list(
            self,
            value,
            template,
            item
    ):

        if not isinstance(
                value,
                list
        ):

            self.fail(
                template,
                item,
                "不是list"
            )

            return False

        if len(value) == 0:

            self.fail(
                template,
                item,
                "为空"
            )

            return False

        self.ok(
            template,
            item
        )

        return True

    def check_url(
            self,
            value,
            template,
            item
    ):

        if not value:

            self.fail(
                template,
                item,
                "URL为空"
            )

            return False

        if not str(value).startswith("http"):

            self.fail(
                template,
                item,
                "不是URL"
            )

            return False

        self.ok(
            template,
            item
        )

        return True


# ==========================================
# JSON解析
# ==========================================

class TemplateParser:

    def __init__(self, filename):

        with open(
                filename,
                "r",
                encoding="utf8"
        ) as f:

            self.data = json.load(f)

        self.templates = self.data["res"]

    def all(self):

        return self.templates

    def get(self, code):

        result = []

        for x in self.templates:

            if x["template_code"] == code:

                result.append(x)

        return result


# ==========================================
# 基类
# ==========================================

class BaseValidator:

    def __init__(
            self,
            checker
    ):

        self.c = checker

    def validate(
            self,
            template
    ):

        pass
        

# ==========================================
# SUMMARY_V1
# ==========================================

class SummaryV1Validator(BaseValidator):

    def validate(self, template):

        name = "SUMMARY_V1"

        data = template.get("data", {})

        knowledges = data.get("knowledges")

        if not knowledges:
            self.c.fail(name, "knowledges", "不存在")
            return

        # ------------------------
        # video
        # ------------------------

        self.c.check_url(
            knowledges.get("video"),
            name,
            "video"
        )

        # ------------------------
        # words
        # ------------------------

        words = knowledges.get("words")

        if self.c.check_list(
                words,
                name,
                "words"):

            for index, word in enumerate(words):

                prefix = f"word[{index}]"

                self.c.check_exist(
                    word.get("knowledge_name"),
                    name,
                    prefix + ".knowledge_name"
                )

                self.c.check_url(
                    word.get("image"),
                    name,
                    prefix + ".image"
                )

                ability = word.get(
                    "ability_level",
                    {}
                )

                if not ability:

                    self.c.fail(
                        name,
                        prefix + ".ability_level",
                        "为空"
                    )

                else:

                    self.c.ok(
                        name,
                        prefix + ".ability_level"
                    )

                    for k, v in ability.items():

                        if v not in [1, 2, 3, 4]:

                            self.c.fail(
                                name,
                                prefix,
                                f"{k}={v} 非法"
                            )

                        else:

                            self.c.ok(
                                name,
                                prefix + "." + k
                            )

        # ------------------------
        # sentence_patterns
        # ------------------------

        patterns = knowledges.get(
            "sentence_patterns"
        )

        if patterns:

            for index, item in enumerate(patterns):

                prefix = f"sentence[{index}]"

                self.c.check_exist(
                    item.get("knowledge_name"),
                    name,
                    prefix + ".knowledge_name"
                )

                ability = item.get(
                    "ability_level",
                    {}
                )

                if not ability:

                    self.c.fail(
                        name,
                        prefix,
                        "ability为空"
                    )

                else:

                    self.c.ok(
                        name,
                        prefix + ".ability"
                    )

                    for k, v in ability.items():

                        if v not in [1, 2, 3, 4]:

                            self.c.fail(
                                name,
                                prefix,
                                f"{k}={v}"
                            )

                        else:

                            self.c.ok(
                                name,
                                prefix + "." + k
                            )

        else:

            self.c.fail(
                name,
                "sentence_patterns",
                "为空"
            )


# ==========================================
# SUMMARY_PERSON
# ==========================================

class SummaryPersonValidator(BaseValidator):

    def validate(self, template):

        name = "SUMMARY_PERSON"

        data = template.get(
            "data",
            {}
        )

        content = data.get(
            "content"
        )

        if not content:

            self.c.fail(
                name,
                "content",
                "不存在"
            )
            return

        self.c.check_exist(
            content.get("title"),
            name,
            "title"
        )

        self.c.check_exist(
            content.get("good_performance_title"),
            name,
            "good_performance_title"
        )

        self.c.check_exist(
            content.get("need_strengthen_title"),
            name,
            "need_strengthen_title"
        )

        # ------------------------
        # 优秀表现
        # ------------------------

        good = content.get(
            "good_performance"
        )

        if self.c.check_list(
                good,
                name,
                "good_performance"):

            for i, text in enumerate(good):

                self.c.check_exist(
                    text,
                    name,
                    f"good[{i}]"
                )

        # ------------------------
        # 加强内容
        # ------------------------

        need = content.get(
            "need_strengthen"
        )

        if self.c.check_list(
                need,
                name,
                "need_strengthen"):

            for i, text in enumerate(need):

                self.c.check_exist(
                    text,
                    name,
                    f"need[{i}]"
                )

        # ------------------------
        # 默认文案
        # ------------------------

        self.c.check_list(
            content.get("good_default"),
            name,
            "good_default"
        )

        self.c.check_list(
            content.get("need_default"),
            name,
            "need_default"
        )

        # ------------------------
        # audio
        # ------------------------

        audio = content.get("audio")

        if audio is None:

            self.c.fail(
                name,
                "audio",
                "不存在"
            )

        else:

            self.c.ok(
                name,
                "audio"
            )
            
# ==========================================
# LISTEN_CHOOSE_PIC
# ==========================================

class ListenChoosePicValidator(BaseValidator):

    def validate(self, template):

        name = "LISTEN_CHOOSE_PIC"

        data = template.get("data", {})
        questions = data.get("questions")

        if not self.c.check_list(
                questions,
                name,
                "questions"):
            return

        for q_index, question in enumerate(questions):

            prefix = f"question[{q_index}]"

            self.c.check_exist(
                question.get("question_id"),
                name,
                prefix + ".question_id"
            )

            title = question.get("title", {})

            self.c.check_exist(
                title.get("text"),
                name,
                prefix + ".title.text"
            )

            self.c.check_url(
                title.get("audio"),
                name,
                prefix + ".title.audio"
            )

            options = question.get("options")

            if not self.c.check_list(
                    options,
                    name,
                    prefix + ".options"):
                continue

            right = 0

            for i, option in enumerate(options):

                p = prefix + f".option[{i}]"

                self.c.check_exist(
                    option.get("knowledge_id"),
                    name,
                    p + ".knowledge_id"
                )

                self.c.check_exist(
                    option.get("text"),
                    name,
                    p + ".text"
                )

                self.c.check_url(
                    option.get("image"),
                    name,
                    p + ".image"
                )

                if option.get("is_right"):
                    right += 1

            if right == 1:

                self.c.ok(
                    name,
                    prefix + ".single_answer"
                )

            else:

                self.c.fail(
                    name,
                    prefix + ".single_answer",
                    f"正确答案数量={right}"
                )


# ==========================================
# MATCH
# ==========================================

class MatchValidator(BaseValidator):

    def validate(self, template):

        name = "MATCH"

        data = template.get("data", {})
        questions = data.get("questions")

        if not self.c.check_list(
                questions,
                name,
                "questions"):
            return

        question = questions[0]

        options = question.get("options")

        if not self.c.check_list(
                options,
                name,
                "options"):
            return

        left = []
        right = []

        for option in options:

            if option["option_group"] == 1:
                left.append(option)

            elif option["option_group"] == 2:
                right.append(option)

        if len(left) == len(right):

            self.c.ok(
                name,
                "左右数量一致"
            )

        else:

            self.c.fail(
                name,
                "左右数量一致",
                f"left={len(left)} right={len(right)}"
            )

        # 左边检查(image)

        for i, item in enumerate(left):

            prefix = f"left[{i}]"

            self.c.check_exist(
                item.get("knowledge_id"),
                name,
                prefix + ".knowledge_id"
            )

            self.c.check_exist(
                item.get("text"),
                name,
                prefix + ".text"
            )

            self.c.check_url(
                item.get("image"),
                name,
                prefix + ".image"
            )

        # 右边检查(audio)

        for i, item in enumerate(right):

            prefix = f"right[{i}]"

            self.c.check_exist(
                item.get("knowledge_id"),
                name,
                prefix + ".knowledge_id"
            )

            self.c.check_exist(
                item.get("text"),
                name,
                prefix + ".text"
            )

            self.c.check_url(
                item.get("audio"),
                name,
                prefix + ".audio"
            )

        # knowledge_id对应

        left_ids = sorted([
            x["knowledge_id"]
            for x in left
        ])

        right_ids = sorted([
            x["knowledge_id"]
            for x in right
        ])

        if left_ids == right_ids:

            self.c.ok(
                name,
                "knowledge_id匹配"
            )

        else:

            self.c.fail(
                name,
                "knowledge_id匹配",
                "左右knowledge_id不一致"
            )


# ==========================================
# PIC_TEXT_CHOOSE_AUDIO
# ==========================================

class PicTextChooseAudioValidator(BaseValidator):

    def validate(self, template):

        name = "PIC_TEXT_CHOOSE_AUDIO"

        data = template.get("data", {})
        questions = data.get("questions")

        if not self.c.check_list(
                questions,
                name,
                "questions"):
            return

        for q_index, question in enumerate(questions):

            prefix = f"question[{q_index}]"

            title = question.get("title", {})

            self.c.check_exist(
                title.get("text"),
                name,
                prefix + ".title.text"
            )

            self.c.check_url(
                title.get("image"),
                name,
                prefix + ".title.image"
            )

            options = question.get("options")

            if not self.c.check_list(
                    options,
                    name,
                    prefix + ".options"):
                continue

            right = 0

            for i, option in enumerate(options):

                p = prefix + f".option[{i}]"

                self.c.check_exist(
                    option.get("text"),
                    name,
                    p + ".text"
                )

                self.c.check_url(
                    option.get("audio"),
                    name,
                    p + ".audio"
                )

                if option.get("is_right"):
                    right += 1

            if right == 1:

                self.c.ok(
                    name,
                    prefix + ".single_answer"
                )

            else:

                self.c.fail(
                    name,
                    prefix + ".single_answer",
                    f"正确答案数量={right}"
                )
                
# ==========================================
# PIC_TEXT_LISTEN_REPEAT
# ==========================================

class PicTextListenRepeatValidator(BaseValidator):

    def validate(self, template):

        name = "PIC_TEXT_LISTEN_REPEAT"

        questions = template.get("data", {}).get("questions")

        if not self.c.check_list(
                questions,
                name,
                "questions"):
            return

        for i, q in enumerate(questions):

            prefix = f"question[{i}]"

            title = q.get("title", {})

            self.c.check_exist(
                title.get("text"),
                name,
                prefix + ".title.text"
            )

            self.c.check_url(
                title.get("image"),
                name,
                prefix + ".title.image"
            )

            self.c.check_url(
                title.get("audio"),
                name,
                prefix + ".title.audio"
            )
            
# ==========================================
# PIC_TEXT_SPEAK
# ==========================================

class PicTextSpeakValidator(BaseValidator):

    def validate(self, template):

        name = "PIC_TEXT_SPEAK"

        questions = template.get("data", {}).get("questions")

        if not self.c.check_list(
                questions,
                name,
                "questions"):
            return

        for i, q in enumerate(questions):

            prefix = f"question[{i}]"

            title = q.get("title", {})

            self.c.check_exist(
                title.get("text"),
                name,
                prefix + ".title.text"
            )

            self.c.check_exist(
                title.get("answer"),
                name,
                prefix + ".answer"
            )

            self.c.check_url(
                title.get("image"),
                name,
                prefix + ".image"
            )

            text_list = title.get("text_list")

            self.c.check_list(
                text_list,
                name,
                prefix + ".text_list"
            )
            
# ==========================================
# PRE_POST_SETTLE
# ==========================================

class PrePostSettleValidator(BaseValidator):

    def validate(self, template):

        name = "PRE_POST_SETTLE"

        if template.get("data") is None:

            self.c.ok(
                name,
                "data为空"
            )

        else:

            self.c.fail(
                name,
                "data为空",
                "应为None"
            )

validator_map = {

    "SUMMARY_V1":
        SummaryV1Validator,

    "SUMMARY_PERSON":
        SummaryPersonValidator,

    "LISTEN_CHOOSE_PIC":
        ListenChoosePicValidator,

    "MATCH":
        MatchValidator,

    "PIC_TEXT_CHOOSE_AUDIO":
        PicTextChooseAudioValidator,

    "PIC_TEXT_LISTEN_REPEAT":
        PicTextListenRepeatValidator,

    "PIC_TEXT_SPEAK":
        PicTextSpeakValidator,

    "PRE_POST_SETTLE":
        PrePostSettleValidator
}

def main():

    if len(sys.argv) != 2:

        print(
            "python check_701_template.py xxx.json"
        )

        return

    filename = sys.argv[1]

    parser = TemplateParser(filename)

    report = ExcelReport()

    checker = Checker(report)

    print("=" * 60)
    print("701模板自动检查")
    print("=" * 60)

    for template in parser.all():

        code = template["template_code"]

        validator_cls = validator_map.get(code)

        if validator_cls is None:

            print(
                "忽略:",
                code
            )

            continue

        validator = validator_cls(checker)

        validator.validate(template)

    report_name = report.save()

    print()

    print("=" * 60)

    print(
        "PASS:",
        report.pass_count
    )

    print(
        "FAIL:",
        report.fail_count
    )

    print()

    print(
        "Excel:",
        report_name
    )

    print("=" * 60)

    if report.fail_count:

        sys.exit(1)

    sys.exit(0)


if __name__ == "__main__":

    main()